"""
Telegram Notification Service
Handles sending notifications to VIP customers and multiple directors.
"""
import logging
import asyncio
import io
from datetime import datetime
from typing import List, Dict, Any, Optional
from aiogram import Bot
from aiogram.types import BufferedInputFile
from aiogram.enums import ParseMode

from config import config
from excel_generator import excel_generator

logger = logging.getLogger(__name__)


class NotificationService:
    """Service for sending Telegram notifications."""
    
    def __init__(self, bot: Bot):
        self.bot = bot
        self.company_name = config.COMPANY_NAME
        self.company_phone = config.COMPANY_PHONE
    
    def _format_money(self, amount: Any) -> str:
        """Format money amount with spaces."""
        if amount is None:
            return "0"
        try:
            num = float(amount)
            return f"{num:,.0f}".replace(",", " ")
        except (ValueError, TypeError):
            return str(amount)
    
    def _get_director_ids(self, director_ids_from_request: List[str] = None) -> List[str]:
        """Get list of director IDs from request or config."""
        # Priority: request > config
        if director_ids_from_request:
            return [id.strip() for id in director_ids_from_request if id and id.strip()]
        return config.get_director_ids()
    
    async def send_purchase_notification(
        self,
        customer_telegram_id: str,
        customer_name: str,
        customer_phone: str,
        sale_number: str,
        sale_date: datetime,
        items: List[Dict[str, Any]],
        total_amount: float,
        paid_amount: float,
        debt_amount: float,
        operator_name: str = "Kassir",
        director_ids: List[str] = None,
        previous_customer_debt: float = 0.0,
        total_customer_debt: float = 0.0
    ) -> Dict[str, Any]:
        """
        Send purchase notification to customer and all directors.
        Directors receive FULL notification (same as customer) + Excel file.
        
        Returns:
            Dict with status and details
        """
        result = {
            "success": False,
            "customer_notified": False,
            "directors_notified": 0,
            "directors_failed": 0,
            "error": None
        }
        
        # Get director IDs
        all_director_ids = self._get_director_ids(director_ids)
        
        # Generate text message for customer
        items_text = ""
        for idx, item in enumerate(items, 1):
            qty = item.get('quantity', 0)
            uom = item.get('uom_symbol', '')
            price = self._format_money(item.get('total_price', 0))
            items_text += f"  {idx}. {item.get('product_name', '')} - {qty} {uom} = {price} so'm\n"
        
        # Build debt section for message
        if debt_amount > 0 and total_customer_debt > 0:
            if previous_customer_debt > 0:
                # Oldingi qarz + yangi qarz = umumiy qarz
                debt_section = (
                    f"⚠️ <b>Bu savdo qarzi:</b> {self._format_money(debt_amount)} so'm\n"
                    f"📊 <b>Umumiy qarzdorlik:</b> {self._format_money(previous_customer_debt)} + "
                    f"{self._format_money(debt_amount)} = <b>{self._format_money(total_customer_debt)}</b> so'm"
                )
            else:
                # Faqat yangi qarz (oldingi qarz yo'q)
                debt_section = f"⚠️ <b>Qarz:</b> {self._format_money(debt_amount)} so'm"
        elif debt_amount > 0:
            debt_section = f"⚠️ <b>Qarz:</b> {self._format_money(debt_amount)} so'm"
        else:
            debt_section = "✅ <b>To'liq to'landi!</b>"
        
        customer_message = f"""
📦 <b>YANGI HARID</b>

👤 <b>Mijoz:</b> {customer_name}
📱 <b>Telefon:</b> {customer_phone}
🧾 <b>Chek:</b> #{sale_number}
📅 <b>Sana:</b> {sale_date.strftime('%d.%m.%Y %H:%M') if isinstance(sale_date, datetime) else sale_date}
👨‍💼 <b>Kassir:</b> {operator_name}

<b>📋 Tovarlar:</b>
{items_text}
💰 <b>Jami:</b> {self._format_money(total_amount)} so'm
✅ <b>To'landi:</b> {self._format_money(paid_amount)} so'm
{debt_section}

━━━━━━━━━━━━━━━━━
🏪 <b>{self.company_name}</b>
📞 {self.company_phone}
Xaridingiz uchun rahmat! 🙏
"""
        
        # Generate Excel file
        excel_file = None
        excel_filename = None
        excel_data = None
        try:
            excel_file = excel_generator.generate_purchase_notification(
                customer_name=customer_name,
                customer_phone=customer_phone,
                sale_number=sale_number,
                sale_date=sale_date,
                items=items,
                total_amount=total_amount,
                paid_amount=paid_amount,
                debt_amount=debt_amount,
                operator_name=operator_name
            )
            excel_filename = f"Harid_{datetime.now().strftime('%Y-%m-%d')}_{sale_number}.xlsx"
            excel_file.seek(0)
            excel_data = excel_file.read()
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
        
        # Send to customer
        if customer_telegram_id:
            try:
                # Send text message
                await self.bot.send_message(
                    chat_id=customer_telegram_id,
                    text=customer_message,
                    parse_mode=ParseMode.HTML
                )
                
                # Send Excel file
                if excel_data:
                    await self.bot.send_document(
                        chat_id=customer_telegram_id,
                        document=BufferedInputFile(
                            file=excel_data,
                            filename=excel_filename
                        ),
                        caption="📊 Harid tafsilotlari"
                    )
                
                result["customer_notified"] = True
                logger.info(f"Purchase notification sent to customer {customer_name} ({customer_telegram_id})")
                
            except Exception as e:
                error_msg = str(e)
                logger.error(f"Failed to send to customer {customer_telegram_id}: {error_msg}")
                result["error"] = error_msg
        else:
            result["error"] = "Customer has no Telegram ID"
            logger.warning(f"Customer {customer_name} has no Telegram ID")
        
        # Send FULL notification to ALL directors (same as customer + additional info)
        if all_director_ids:
            # Message for director with additional info
            director_message = f"""
📊 <b>YANGI HARID - HISOBOT</b>

👤 <b>Mijoz:</b> {customer_name}
📱 <b>Telefon:</b> {customer_phone}
🆔 <b>Telegram:</b> {customer_telegram_id or "Yo'q"}
🧾 <b>Chek:</b> #{sale_number}
📅 <b>Sana:</b> {sale_date.strftime('%d.%m.%Y %H:%M') if isinstance(sale_date, datetime) else sale_date}
👨‍💼 <b>Kassir:</b> {operator_name}

<b>📋 Tovarlar:</b>
{items_text}
💰 <b>Jami:</b> {self._format_money(total_amount)} so'm
✅ <b>To'landi:</b> {self._format_money(paid_amount)} so'm
{debt_section}

📤 <b>Mijozga xabar:</b> {"✅ Yuborildi" if result["customer_notified"] else "❌ Yuborilmadi"}
{f"⚠️ Xato: {result['error']}" if result.get('error') and not result["customer_notified"] else ""}

━━━━━━━━━━━━━━━━━
🏪 <b>{self.company_name}</b>
"""
            
            for director_id in all_director_ids:
                try:
                    # Send text message
                    await self.bot.send_message(
                        chat_id=director_id,
                        text=director_message,
                        parse_mode=ParseMode.HTML
                    )
                    
                    # Send Excel file
                    if excel_data:
                        await self.bot.send_document(
                            chat_id=director_id,
                            document=BufferedInputFile(
                                file=excel_data,
                                filename=excel_filename
                            ),
                            caption="📊 Harid tafsilotlari (Excel)"
                        )
                    
                    result["directors_notified"] += 1
                    logger.info(f"Full notification sent to director {director_id}")
                    
                except Exception as e:
                    result["directors_failed"] += 1
                    logger.error(f"Failed to send to director {director_id}: {e}")
        else:
            logger.warning("No director IDs configured")
        
        result["success"] = result["customer_notified"] or result["directors_notified"] > 0
        return result
    
    async def send_payment_notification(
        self,
        customer_telegram_id: str,
        customer_name: str,
        customer_phone: str,
        payment_date: datetime,
        payment_amount: float,
        payment_type: str,
        previous_debt: float,
        current_debt: float,
        operator_name: str = "Kassir",
        director_ids: List[str] = None
    ) -> Dict[str, Any]:
        """
        Send payment notification to customer and all directors.
        Directors receive FULL notification (same as customer) + Excel file.
        
        Returns:
            Dict with status and details
        """
        result = {
            "success": False,
            "customer_notified": False,
            "directors_notified": 0,
            "directors_failed": 0,
            "error": None
        }
        
        # Get director IDs
        all_director_ids = self._get_director_ids(director_ids)
        
        # Payment type labels
        payment_type_labels = {
            'CASH': '💵 Naqd pul',
            'CARD': '💳 Plastik karta',
            'TRANSFER': '🏦 Bank o\'tkazmasi',
            'MIXED': '💱 Aralash'
        }
        payment_label = payment_type_labels.get(payment_type, payment_type)
        
        # Generate text message
        customer_message = f"""
💰 <b>TO'LOV QABUL QILINDI</b>

👤 <b>Mijoz:</b> {customer_name}
📱 <b>Telefon:</b> {customer_phone}
📅 <b>Sana:</b> {payment_date.strftime('%d.%m.%Y %H:%M') if isinstance(payment_date, datetime) else payment_date}
👨‍💼 <b>Qabul qildi:</b> {operator_name}

{payment_label}

💵 <b>Oldingi qarz:</b> {self._format_money(previous_debt)} so'm
✅ <b>To'landi:</b> {self._format_money(payment_amount)} so'm
{"✅ <b>Qarz to'liq to'landi!</b>" if current_debt <= 0 else f"⚠️ <b>Qolgan qarz:</b> {self._format_money(current_debt)} so'm"}

━━━━━━━━━━━━━━━━━
🏪 <b>{self.company_name}</b>
📞 {self.company_phone}
Rahmat! 🙏
"""
        
        # Generate Excel file
        excel_file = None
        excel_filename = None
        excel_data = None
        try:
            excel_file = excel_generator.generate_payment_notification(
                customer_name=customer_name,
                customer_phone=customer_phone,
                payment_date=payment_date,
                payment_amount=payment_amount,
                payment_type=payment_type,
                previous_debt=previous_debt,
                current_debt=current_debt,
                operator_name=operator_name
            )
            excel_filename = f"Tolov_{datetime.now().strftime('%Y-%m-%d_%H%M')}_{customer_name.replace(' ', '_')}.xlsx"
            excel_file.seek(0)
            excel_data = excel_file.read()
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
        
        # Send to customer
        if customer_telegram_id:
            try:
                # Send text message
                await self.bot.send_message(
                    chat_id=customer_telegram_id,
                    text=customer_message,
                    parse_mode=ParseMode.HTML
                )
                
                # Send Excel file
                if excel_data:
                    await self.bot.send_document(
                        chat_id=customer_telegram_id,
                        document=BufferedInputFile(
                            file=excel_data,
                            filename=excel_filename
                        ),
                        caption="📊 To'lov tafsilotlari"
                    )
                
                result["customer_notified"] = True
                logger.info(f"Payment notification sent to customer {customer_name} ({customer_telegram_id})")
                
            except Exception as e:
                error_msg = str(e)
                logger.error(f"Failed to send to customer {customer_telegram_id}: {error_msg}")
                result["error"] = error_msg
        else:
            result["error"] = "Customer has no Telegram ID"
            logger.warning(f"Customer {customer_name} has no Telegram ID")
        
        # Send FULL notification to ALL directors
        if all_director_ids:
            director_message = f"""
💰 <b>TO'LOV HISOBOTI</b>

👤 <b>Mijoz:</b> {customer_name}
📱 <b>Telefon:</b> {customer_phone}
🆔 <b>Telegram:</b> {customer_telegram_id or "Yo'q"}
📅 <b>Sana:</b> {payment_date.strftime('%d.%m.%Y %H:%M') if isinstance(payment_date, datetime) else payment_date}
👨‍💼 <b>Qabul qildi:</b> {operator_name}

{payment_label}

💵 <b>Oldingi qarz:</b> {self._format_money(previous_debt)} so'm
✅ <b>To'landi:</b> {self._format_money(payment_amount)} so'm
{"✅ <b>Qarz to'liq to'landi!</b>" if current_debt <= 0 else f"⚠️ <b>Qolgan qarz:</b> {self._format_money(current_debt)} so'm"}

📤 <b>Mijozga xabar:</b> {"✅ Yuborildi" if result["customer_notified"] else "❌ Yuborilmadi"}
{f"⚠️ Xato: {result['error']}" if result.get('error') and not result["customer_notified"] else ""}

━━━━━━━━━━━━━━━━━
🏪 <b>{self.company_name}</b>
"""
            
            for director_id in all_director_ids:
                try:
                    # Send text message
                    await self.bot.send_message(
                        chat_id=director_id,
                        text=director_message,
                        parse_mode=ParseMode.HTML
                    )
                    
                    # Send Excel file
                    if excel_data:
                        await self.bot.send_document(
                            chat_id=director_id,
                            document=BufferedInputFile(
                                file=excel_data,
                                filename=excel_filename
                            ),
                            caption="📊 To'lov tafsilotlari (Excel)"
                        )
                    
                    result["directors_notified"] += 1
                    logger.info(f"Full notification sent to director {director_id}")
                    
                except Exception as e:
                    result["directors_failed"] += 1
                    logger.error(f"Failed to send to director {director_id}: {e}")
        else:
            logger.warning("No director IDs configured")
        
        result["success"] = result["customer_notified"] or result["directors_notified"] > 0
        return result
    
    async def send_test_message(self, chat_id: str, message: str = "Test message") -> bool:
        """Send a test message to verify bot is working."""
        try:
            await self.bot.send_message(chat_id=chat_id, text=message)
            return True
        except Exception as e:
            logger.error(f"Test message failed: {e}")
            return False
    
    async def send_daily_report(self, chat_id: str, message: str) -> bool:
        """
        Send daily report to Telegram group.
        
        Args:
            chat_id: Group chat ID (starts with - for groups)
            message: HTML formatted report message
            
        Returns:
            True if sent successfully
        """
        try:
            await self.bot.send_message(
                chat_id=chat_id,
                text=message,
                parse_mode=ParseMode.HTML
            )
            logger.info(f"Daily report sent to group {chat_id}")
            return True
        except Exception as e:
            logger.error(f"Failed to send daily report to {chat_id}: {e}")
            return False
    
    async def send_daily_report_with_excel(self, chat_id: str, report_data: dict) -> bool:
        """
        Send daily report with Excel file to Telegram group.
        
        Args:
            chat_id: Group chat ID
            report_data: Dict containing all report data
            
        Returns:
            True if sent successfully
        """
        try:
            from datetime import date
            import io
            
            # Format money helper
            def fmt(amount):
                return f"{float(amount or 0):,.0f}".replace(",", " ")
            
            today = date.today()
            
            # Build text message
            total_sales = report_data.get("total_sales_count", 0)
            total_amount = report_data.get("total_amount", 0)
            total_paid = report_data.get("total_paid", 0)
            total_debt = report_data.get("total_debt", 0)
            total_discount = report_data.get("total_discount", 0)
            cash_amount = report_data.get("cash_amount", 0)
            card_amount = report_data.get("card_amount", 0)
            transfer_amount = report_data.get("transfer_amount", 0)
            total_all_debt = report_data.get("total_all_debt", 0)
            total_cost = report_data.get("total_cost", 0)
            total_profit = report_data.get("total_profit", 0)
            
            cashiers = report_data.get("cashiers", [])
            debtors = report_data.get("debtors", [])
            products = report_data.get("products", [])
            low_stock = report_data.get("low_stock", [])
            
            lines = [
                f"📊 <b>KUNLIK HISOBOT</b>",
                f"📅 <b>{today.strftime('%d.%m.%Y')}</b>",
                "",
                "═══════════════════════════",
                f"📦 <b>SOTUVLAR</b>",
                "═══════════════════════════",
                f"🛒 Sotuvlar: <b>{total_sales} ta</b>",
                f"💰 Jami: <b>{fmt(total_amount)} so'm</b>",
                f"✅ To'langan: <b>{fmt(total_paid)} so'm</b>",
                f"🔴 Qarz: <b>{fmt(total_debt)} so'm</b>",
            ]
            
            if total_discount > 0:
                lines.append(f"🏷 Chegirma: <b>{fmt(total_discount)} so'm</b>")
            
            lines.extend([
                "",
                "═══════════════════════════",
                f"💳 <b>TO'LOV TURLARI</b>",
                "═══════════════════════════",
                f"💵 Naqd: <b>{fmt(cash_amount)} so'm</b>",
                f"💳 Plastik: <b>{fmt(card_amount)} so'm</b>",
                f"🏦 O'tkazma: <b>{fmt(transfer_amount)} so'm</b>",
                "",
                "═══════════════════════════",
                f"📈 <b>YALPI FOYDA</b>",
                "═══════════════════════════",
                f"📦 Kelish narxi: <b>{fmt(total_cost)} so'm</b>",
                f"💰 Sotuv summasi: <b>{fmt(total_amount)} so'm</b>",
                f"📈 Yalpi foyda: <b>{fmt(total_profit)} so'm</b>",
            ])

            # Expenses section
            total_expenses = report_data.get("total_expenses", 0)
            net_profit_val = report_data.get("net_profit", total_profit - float(total_expenses or 0))
            expenses_list = report_data.get("expenses_list", [])

            if float(total_expenses or 0) > 0:
                lines.extend([
                    "",
                    "═══════════════════════════",
                    f"💸 <b>CHIQIMLAR</b>",
                    "═══════════════════════════",
                ])
                for exp in expenses_list:
                    lines.append(f"  • {exp.get('name', 'Boshqa')}: <b>{fmt(exp.get('total', 0))} so'm</b>")
                lines.append(f"📊 Jami chiqim: <b>{fmt(total_expenses)} so'm</b>")

            lines.extend([
                "",
                "═══════════════════════════",
                f"💰 <b>SOF FOYDA</b>",
                "═══════════════════════════",
                f"📈 Yalpi foyda: <b>{fmt(total_profit)} so'm</b>",
                f"💸 Chiqimlar: <b>{fmt(total_expenses)} so'm</b>",
                f"{'✅' if float(net_profit_val or 0) >= 0 else '🔴'} Sof foyda: <b>{fmt(net_profit_val)} so'm</b>",
            ])
            
            # Cashiers summary
            if cashiers:
                lines.extend([
                    "",
                    "═══════════════════════════",
                    f"👥 <b>KASSIRLAR</b>",
                    "═══════════════════════════",
                ])
                for c in cashiers[:5]:
                    lines.append(f"• {c['name']}: {c['sales_count']} ta / {fmt(c['total_amount'])} so'm")
            
            # Total debt
            lines.extend([
                "",
                f"📊 <b>Jami qarzdorlik:</b> {fmt(total_all_debt)} so'm",
            ])
            
            # Low stock warning
            if low_stock:
                lines.extend([
                    "",
                    f"⚠️ <b>Kam qolgan:</b> {len(low_stock)} ta tovar",
                ])

            # Suppliers debt section
            total_supplier_debt = report_data.get("total_supplier_debt", 0)
            suppliers_list = report_data.get("suppliers_list", [])
            if float(total_supplier_debt or 0) > 0:
                lines.extend([
                    "",
                    "═══════════════════════════",
                    f"🚛 <b>TA'MINOTCHILAR QARZI</b>",
                    "═══════════════════════════",
                ])
                for sup in suppliers_list[:10]:
                    lines.append(f"  • {sup.get('name', '?')}: <b>{fmt(sup.get('debt', 0))} so'm</b>")
                lines.append(f"📊 Jami ta'minotchilarga qarz: <b>{fmt(total_supplier_debt)} so'm</b>")
            
            lines.extend([
                "",
                "═══════════════════════════",
                f"🕐 {datetime.now().strftime('%H:%M')}",
                "═══════════════════════════",
                "",
                "📎 <i>Batafsil ma'lumot Excel faylda</i>",
                "",
                "🏭 <i>VEGAS</i>"
            ])
            
            message = "\n".join(lines)
            
            # Send text message first
            await self.bot.send_message(
                chat_id=chat_id,
                text=message,
                parse_mode=ParseMode.HTML
            )
            
            # Generate Excel file
            excel_data = await self._generate_daily_report_excel(report_data, today)
            
            if excel_data:
                # Send Excel file
                await self.bot.send_document(
                    chat_id=chat_id,
                    document=BufferedInputFile(
                        file=excel_data,
                        filename=f"kunlik_hisobot_{today.strftime('%Y-%m-%d')}.xlsx"
                    ),
                    caption=f"📊 Kunlik hisobot - {today.strftime('%d.%m.%Y')}"
                )
            
            logger.info(f"Daily report with Excel sent to group {chat_id}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to send daily report with Excel: {e}")
            return False
    
    async def _generate_daily_report_excel(self, data: dict, report_date) -> bytes:
        """Generate Excel file for daily report."""
        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            money_font = Font(bold=True, color="228B22")
            debt_font = Font(bold=True, color="DC143C")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            def fmt_money(val):
                return float(val or 0)
            
            # ===== Sheet 1: Umumiy =====
            ws1 = wb.active
            ws1.title = "Umumiy"
            
            ws1['A1'] = "KUNLIK HISOBOT"
            ws1['A1'].font = Font(bold=True, size=16)
            ws1['A2'] = f"Sana: {report_date.strftime('%d.%m.%Y')}"
            
            summary_data = [
                ["Ko'rsatkich", "Qiymat"],
                ["Sotuvlar soni", f"{data.get('total_sales_count', 0)} ta"],
                ["Jami summa", f"{fmt_money(data.get('total_amount', 0)):,.0f} so'm"],
                ["To'langan", f"{fmt_money(data.get('total_paid', 0)):,.0f} so'm"],
                ["Qarz", f"{fmt_money(data.get('total_debt', 0)):,.0f} so'm"],
                ["Chegirma", f"{fmt_money(data.get('total_discount', 0)):,.0f} so'm"],
                ["", ""],
                ["TO'LOV TURLARI", ""],
                ["Naqd pul", f"{fmt_money(data.get('cash_amount', 0)):,.0f} so'm"],
                ["Plastik karta", f"{fmt_money(data.get('card_amount', 0)):,.0f} so'm"],
                ["O'tkazma", f"{fmt_money(data.get('transfer_amount', 0)):,.0f} so'm"],
                ["", ""],
                ["YALPI FOYDA", ""],
                ["Kelish narxi (tan narxi)", f"{fmt_money(data.get('total_cost', 0)):,.0f} so'm"],
                ["Sotuv summasi", f"{fmt_money(data.get('total_amount', 0)):,.0f} so'm"],
                ["Yalpi foyda", f"{fmt_money(data.get('total_profit', 0)):,.0f} so'm"],
                ["", ""],
                ["CHIQIMLAR", ""],
            ]

            # Add expense categories
            for exp in data.get('expenses_list', []):
                summary_data.append([exp.get('name', 'Boshqa'), f"{fmt_money(exp.get('total', 0)):,.0f} so'm"])
            summary_data.append(["Jami chiqimlar", f"{fmt_money(data.get('total_expenses', 0)):,.0f} so'm"])

            summary_data.extend([
                ["", ""],
                ["SOF FOYDA", ""],
                ["Yalpi foyda", f"{fmt_money(data.get('total_profit', 0)):,.0f} so'm"],
                ["Chiqimlar", f"{fmt_money(data.get('total_expenses', 0)):,.0f} so'm"],
                ["Sof foyda", f"{fmt_money(data.get('net_profit', 0)):,.0f} so'm"],
                ["", ""],
                ["Umumiy qarzdorlik", f"{fmt_money(data.get('total_all_debt', 0)):,.0f} so'm"],
            ])
            
            profit_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            profit_font = Font(bold=True, color="228B22", size=12)

            for row_idx, row_data in enumerate(summary_data, start=4):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    if row_idx == 4:
                        cell.font = header_font
                        cell.fill = header_fill
                # Highlight "Sof foyda" row (19 = 4 + 15, index 15 in summary_data)
                if row_data[0] == "Sof foyda":
                    for col_idx in range(1, 3):
                        cell = ws1.cell(row=row_idx, column=col_idx)
                        cell.font = profit_font
                        cell.fill = profit_fill
            
            ws1.column_dimensions['A'].width = 25
            ws1.column_dimensions['B'].width = 30
            
            # ===== Sheet 2: Kassirlar =====
            ws2 = wb.create_sheet("Kassirlar")
            
            cashier_headers = ["№", "Kassir", "Sotuvlar", "Jami summa", "To'langan", "Qarz"]
            for col, header in enumerate(cashier_headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            for idx, c in enumerate(data.get('cashiers', []), 1):
                ws2.cell(row=idx+1, column=1, value=idx).border = border
                ws2.cell(row=idx+1, column=2, value=c.get('name', '')).border = border
                ws2.cell(row=idx+1, column=3, value=c.get('sales_count', 0)).border = border
                ws2.cell(row=idx+1, column=4, value=fmt_money(c.get('total_amount', 0))).border = border
                ws2.cell(row=idx+1, column=5, value=fmt_money(c.get('paid_amount', 0))).border = border
                ws2.cell(row=idx+1, column=6, value=fmt_money(c.get('debt_amount', 0))).border = border
            
            for col in range(1, 7):
                ws2.column_dimensions[get_column_letter(col)].width = 18
            
            # ===== Sheet 3: Sotilgan tovarlar =====
            ws3 = wb.create_sheet("Sotilgan tovarlar")
            
            product_headers = ["№", "Mahsulot", "Miqdor", "O'lchov", "Summa"]
            for col, header in enumerate(product_headers, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            for idx, p in enumerate(data.get('products', []), 1):
                ws3.cell(row=idx+1, column=1, value=idx).border = border
                ws3.cell(row=idx+1, column=2, value=p.get('name', '')).border = border
                ws3.cell(row=idx+1, column=3, value=float(p.get('quantity', 0))).border = border
                ws3.cell(row=idx+1, column=4, value=p.get('uom', '')).border = border
                ws3.cell(row=idx+1, column=5, value=fmt_money(p.get('total', 0))).border = border
            
            ws3.column_dimensions['A'].width = 8
            ws3.column_dimensions['B'].width = 40
            ws3.column_dimensions['C'].width = 15
            ws3.column_dimensions['D'].width = 12
            ws3.column_dimensions['E'].width = 20
            
            # ===== Sheet 4: Qarzdorlar =====
            ws4 = wb.create_sheet("Bugungi qarzdorlar")
            ws4.sheet_properties.tabColor = "FF0000"
            
            debtor_headers = ["№", "Mijoz", "Telefon", "Qarz summasi"]
            for col, header in enumerate(debtor_headers, 1):
                cell = ws4.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                cell.border = border
            
            for idx, d in enumerate(data.get('debtors', []), 1):
                ws4.cell(row=idx+1, column=1, value=idx).border = border
                ws4.cell(row=idx+1, column=2, value=d.get('name', '')).border = border
                ws4.cell(row=idx+1, column=3, value=d.get('phone', '')).border = border
                cell = ws4.cell(row=idx+1, column=4, value=fmt_money(d.get('debt_amount', 0)))
                cell.border = border
                cell.font = debt_font
            
            for col in range(1, 5):
                ws4.column_dimensions[get_column_letter(col)].width = 20
            
            # ===== Sheet 5: Kam qolgan tovarlar =====
            ws5 = wb.create_sheet("Kam qolgan tovarlar")
            ws5.sheet_properties.tabColor = "FFA500"
            
            low_headers = ["№", "Mahsulot", "Qoldiq", "O'lchov"]
            for col, header in enumerate(low_headers, 1):
                cell = ws5.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                cell.border = border
            
            for idx, item in enumerate(data.get('low_stock', []), 1):
                ws5.cell(row=idx+1, column=1, value=idx).border = border
                ws5.cell(row=idx+1, column=2, value=item.get('name', '')).border = border
                ws5.cell(row=idx+1, column=3, value=float(item.get('quantity', 0))).border = border
                ws5.cell(row=idx+1, column=4, value=item.get('uom', '')).border = border
            
            for col in range(1, 5):
                ws5.column_dimensions[get_column_letter(col)].width = 20
            
            # ===== Sheet 6: BARCHA QARZDORLAR (umumiy ro'yxat) =====
            all_debtors = data.get('all_debtors', [])
            if all_debtors:
                ws6 = wb.create_sheet("BARCHA QARZDORLAR")
                ws6.sheet_properties.tabColor = "8B0000"

                # Header
                ws6.merge_cells('A1:F1')
                ws6['A1'] = f"BARCHA QARZDORLAR RO'YXATI - {report_date.strftime('%d.%m.%Y')}"
                ws6['A1'].font = Font(bold=True, size=14)
                ws6['A1'].alignment = Alignment(horizontal='center')

                ws6.merge_cells('A2:F2')
                ws6['A2'] = f"Jami qarzdorlar soni: {len(all_debtors)} | Umumiy qarz: {fmt_money(sum(d.get('total_debt', 0) for d in all_debtors)):,.0f} so'm"
                ws6['A2'].font = Font(bold=True, size=11, color="DC143C")
                ws6['A2'].alignment = Alignment(horizontal='center')

                # Table headers
                all_debt_headers = ["№", "Mijoz ismi", "Telefon", "Umumiy qarz", "Qarz savdolar", "Oxirgi to'lov"]
                for col, header in enumerate(all_debt_headers, 1):
                    cell = ws6.cell(row=4, column=col, value=header)
                    cell.font = header_font
                    cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

                for idx, debtor in enumerate(all_debtors, 1):
                    row = idx + 4
                    ws6.cell(row=row, column=1, value=idx).border = border
                    ws6.cell(row=row, column=2, value=debtor.get('name', '')).border = border
                    ws6.cell(row=row, column=3, value=debtor.get('phone', '')).border = border

                    debt_cell = ws6.cell(row=row, column=4, value=fmt_money(debtor.get('total_debt', 0)))
                    debt_cell.border = border
                    debt_cell.font = debt_font
                    debt_cell.number_format = '#,##0'

                    ws6.cell(row=row, column=5, value=debtor.get('sales_count', 0)).border = border

                    # Last payment
                    payments = debtor.get('payments', [])
                    last_payment = payments[0] if payments else None
                    if last_payment:
                        ws6.cell(row=row, column=6, value=f"{last_payment.get('date', '')} - {fmt_money(last_payment.get('amount', 0)):,.0f}").border = border
                    else:
                        ws6.cell(row=row, column=6, value="To'lov yo'q").border = border

                ws6.column_dimensions['A'].width = 6
                ws6.column_dimensions['B'].width = 30
                ws6.column_dimensions['C'].width = 18
                ws6.column_dimensions['D'].width = 20
                ws6.column_dimensions['E'].width = 15
                ws6.column_dimensions['F'].width = 25

                # ===== Sheet 7: QARZDORLAR TAFSILOTI =====
                ws7 = wb.create_sheet("Qarzdorlar tafsiloti")
                ws7.sheet_properties.tabColor = "8B0000"

                current_row = 1

                for debtor in all_debtors[:50]:  # Top 50 debtors with details
                    # Debtor header
                    ws7.merge_cells(f'A{current_row}:G{current_row}')
                    header_cell = ws7.cell(row=current_row, column=1,
                        value=f"👤 {debtor.get('name', '')} | 📱 {debtor.get('phone', '')} | 💰 Qarz: {fmt_money(debtor.get('total_debt', 0)):,.0f} so'm")
                    header_cell.font = Font(bold=True, size=12, color="FFFFFF")
                    header_cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
                    header_cell.alignment = Alignment(horizontal='left')
                    current_row += 1

                    # Sales section
                    sales = debtor.get('sales', [])
                    if sales:
                        ws7.cell(row=current_row, column=1, value="📋 QARZ SAVDOLARI:").font = Font(bold=True, size=10)
                        current_row += 1

                        # Sales headers
                        sale_headers = ["Sana", "Chek №", "Tovarlar", "Jami", "To'langan", "Qarz"]
                        for col, h in enumerate(sale_headers, 1):
                            cell = ws7.cell(row=current_row, column=col, value=h)
                            cell.font = Font(bold=True, size=9)
                            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            cell.border = border
                        current_row += 1

                        for sale in sales[:10]:  # Last 10 sales
                            # Items summary
                            items_text = ", ".join([f"{i.get('product', '')[:20]} ({i.get('quantity', 0)} {i.get('uom', '')})"
                                                   for i in sale.get('items', [])[:3]])
                            if len(sale.get('items', [])) > 3:
                                items_text += f" +{len(sale.get('items', [])) - 3} ta"

                            ws7.cell(row=current_row, column=1, value=f"{sale.get('date', '')} {sale.get('time', '')}").border = border
                            ws7.cell(row=current_row, column=2, value=sale.get('sale_number', '')).border = border
                            ws7.cell(row=current_row, column=3, value=items_text[:50]).border = border
                            ws7.cell(row=current_row, column=4, value=fmt_money(sale.get('total_amount', 0))).border = border
                            ws7.cell(row=current_row, column=5, value=fmt_money(sale.get('paid_amount', 0))).border = border

                            debt_cell = ws7.cell(row=current_row, column=6, value=fmt_money(sale.get('debt_amount', 0)))
                            debt_cell.border = border
                            debt_cell.font = debt_font
                            current_row += 1

                    # Payments section
                    payments = debtor.get('payments', [])
                    if payments:
                        current_row += 1
                        ws7.cell(row=current_row, column=1, value="💵 TO'LOVLAR TARIXI:").font = Font(bold=True, size=10)
                        current_row += 1

                        payment_headers = ["Sana", "Summa", "To'lov turi", "Izoh"]
                        for col, h in enumerate(payment_headers, 1):
                            cell = ws7.cell(row=current_row, column=col, value=h)
                            cell.font = Font(bold=True, size=9)
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                            cell.border = border
                        current_row += 1

                        payment_types = {'cash': 'Naqd', 'card': 'Karta', 'transfer': "O'tkazma"}
                        for payment in payments[:5]:  # Last 5 payments
                            ws7.cell(row=current_row, column=1, value=payment.get('date', '')).border = border

                            amount_cell = ws7.cell(row=current_row, column=2, value=fmt_money(payment.get('amount', 0)))
                            amount_cell.border = border
                            amount_cell.font = money_font

                            ws7.cell(row=current_row, column=3, value=payment_types.get(payment.get('payment_type', ''), payment.get('payment_type', ''))).border = border
                            ws7.cell(row=current_row, column=4, value=payment.get('notes', '')[:30]).border = border
                            current_row += 1

                    # Empty row between debtors
                    current_row += 2

                # Set column widths
                ws7.column_dimensions['A'].width = 18
                ws7.column_dimensions['B'].width = 15
                ws7.column_dimensions['C'].width = 45
                ws7.column_dimensions['D'].width = 15
                ws7.column_dimensions['E'].width = 15
                ws7.column_dimensions['F'].width = 15
                ws7.column_dimensions['G'].width = 20

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except Exception as e:
            logger.error(f"Failed to generate Excel: {e}")
            return None