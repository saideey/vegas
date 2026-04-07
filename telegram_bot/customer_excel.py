"""
Customer Excel Report Generator for Telegram Bot.
Generates the same report as the frontend Excel export.
"""
import io
from datetime import datetime
from typing import Dict, Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def fmt_money(amount) -> str:
    try:
        n = float(amount)
        return f"{n:,.0f}".replace(",", " ")
    except:
        return "0"


def fmt_usd(amount) -> str:
    try:
        n = float(amount)
        return f"${n:,.2f}"
    except:
        return "$0.00"


def fmt_date(iso_str: str) -> str:
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str)
        return dt.strftime("%d.%m.%Y")
    except:
        return iso_str[:10] if len(iso_str) >= 10 else iso_str


def fmt_time(iso_str: str) -> str:
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str)
        return dt.strftime("%H:%M")
    except:
        return ""


def generate_customer_excel(data: Dict[str, Any]) -> bytes:
    """Generate Excel report bytes for customer."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl not installed")

    customer = data.get("customer", {})
    sales = data.get("sales", [])
    debt_history = data.get("debt_history", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Mijoz hisoboti"

    # Colors
    C = {
        "header":    "1F4E79",
        "subheader": "2E75B6",
        "success":   "375623",
        "danger":    "C00000",
        "warning":   "FF8C00",
        "white":     "FFFFFF",
        "lightBlue": "DEEAF1",
        "lightRed":  "FCE4D6",
        "lightGreen":"E2EFDA",
        "gray":      "808080",
        "yellow":    "FFF2CC",
        "blue_dark": "1E40AF",
        "blue_light":"DBEAFE",
    }

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def fill(color): return PatternFill("solid", fgColor="FF" + color)
    def font(color="000000", bold=False, size=11, italic=False):
        return Font(bold=bold, size=size, color="FF" + color, italic=italic)
    def align(h="left", v="middle", wrap=False):
        # openpyxl vertical: top, center, bottom, distributed, justify
    v_map = {"middle": "center"}
    v = v_map.get(v, v)
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    row = 1

    # ── SECTION 1: Header ──
    ws.merge_cells(f"A{row}:K{row}")
    c = ws.cell(row, 1)
    c.value = f"🏢 Vegas - MIJOZ HISOBOTI"
    c.font = Font(bold=True, size=18, color="FFFFFFFF")
    c.fill = fill(C["header"])
    c.alignment = align("center")
    ws.row_dimensions[row].height = 35
    row += 2

    # ── SECTION 2: Customer info ──
    ws.merge_cells(f"A{row}:E{row}")
    h1 = ws.cell(row, 1)
    h1.value = "👤 MIJOZ MA'LUMOTLARI"
    h1.font = font(C["white"], bold=True, size=12)
    h1.fill = fill(C["subheader"])
    h1.alignment = align("center")

    ws.merge_cells(f"F{row}:K{row}")
    h2 = ws.cell(row, 6)
    h2.value = "📊 MOLIYAVIY STATISTIKA"
    h2.font = font(C["white"], bold=True, size=12)
    h2.fill = fill(C["success"])
    h2.alignment = align("center")
    row += 1

    # Calculate stats
    total_sales_count = len(sales)
    total_amount = sum(float(s.get("total_amount", 0)) for s in sales)
    total_paid = sum(float(s.get("paid_amount", 0)) for s in sales)
    total_debt_from_sales = sum(float(s.get("debt_amount", 0)) for s in sales)
    uzs_payments = sum(
        float(r.get("amount", 0)) for r in debt_history
        if (r.get("transaction_type") in ("PAYMENT", "DEBT_PAYMENT")) and r.get("currency", "UZS") == "UZS"
    )
    total_items = sum(len(s.get("items", [])) for s in sales)

    debt_uzs = float(customer.get("current_debt", 0))
    debt_usd = float(customer.get("current_debt_usd", 0))
    advance = float(customer.get("advance_balance", 0))

    info_rows = [
        ["Mijoz ismi:", customer.get("name", ""), "", "", "", "Jami xaridlar:", f"{total_sales_count} ta"],
        ["Telefon:", customer.get("phone", ""), "", "", "", "Jami summa:", fmt_money(total_amount)],
        ["Kompaniya:", customer.get("company_name", "") or "-", "", "", "", "To'langan:", fmt_money(total_paid)],
        ["Manzil:", customer.get("address", "") or "-", "", "", "", "Qarz (sotuvlardan):", fmt_money(total_debt_from_sales)],
        ["Mijoz turi:", customer.get("customer_type", ""), "", "", "", "To'lovlar (alohida):", fmt_money(uzs_payments)],
        ["", "", "", "", "", "Sotib olingan tovarlar:", f"{total_items} ta"],
        ["", "", "", "", "", "Joriy qarz (so'm):", fmt_money(debt_uzs) + " so'm" if debt_uzs > 0 else "0 so'm"],
        ["", "", "", "", "", "Joriy qarz ($):", fmt_usd(debt_usd)],
        ["", "", "", "", "", "Avans balansi:", fmt_money(advance) + " so'm"],
    ]

    for i, info_row in enumerate(info_rows):
        r = ws.append(info_row)
        cur = ws._current_row
        ws.cell(cur, 1).font = font(C["subheader"], bold=True)
        ws.cell(cur, 1).fill = fill(C["lightBlue"])
        ws.cell(cur, 2).fill = fill(C["white"])
        # Color stats column
        if i == 6:  # UZS debt
            ws.cell(cur, 6).font = font(C["danger"] if debt_uzs > 0 else C["success"], bold=True)
            ws.cell(cur, 6).fill = fill(C["lightBlue"])
            ws.cell(cur, 7).font = font(C["danger"] if debt_uzs > 0 else C["success"], bold=True)
            ws.cell(cur, 7).fill = fill(C["lightRed"] if debt_uzs > 0 else C["lightGreen"])
        elif i == 7:  # USD debt
            ws.cell(cur, 6).font = font(C["blue_dark"], bold=True)
            ws.cell(cur, 6).fill = fill(C["blue_light"])
            ws.cell(cur, 7).font = font(C["blue_dark"] if debt_usd > 0 else C["success"], bold=True)
            ws.cell(cur, 7).fill = fill(C["blue_light"] if debt_usd > 0 else C["lightGreen"])
        elif i == 8:  # Advance
            ws.cell(cur, 6).font = font(C["success"], bold=True)
            ws.cell(cur, 6).fill = fill(C["lightGreen"])
            ws.cell(cur, 7).font = font(C["success"], bold=True)
            ws.cell(cur, 7).fill = fill(C["lightGreen"])
        else:
            ws.cell(cur, 6).font = font(C["success"], bold=True)
            ws.cell(cur, 6).fill = fill(C["lightGreen"])
            ws.cell(cur, 7).alignment = align("right")
        row += 1

    row += 1

    # ── SECTION 3: Sales ──
    ws.merge_cells(f"A{row}:K{row}")
    c = ws.cell(row, 1)
    c.value = f"🛒 SOTUVLAR TARIXI ({total_sales_count} ta xarid)"
    c.font = font(C["white"], bold=True, size=12)
    c.fill = fill(C["warning"])
    c.alignment = align("center")
    ws.row_dimensions[row].height = 22
    row += 1

    sale_headers = ["№", "Sana", "Chek №", "Tovarlar soni", "Umumiy summa", "To'langan", "Qarz", "Holat", "", "", ""]
    hr = ws.append(sale_headers)
    cur = ws._current_row
    for col in range(1, 9):
        cell = ws.cell(cur, col)
        cell.font = font(C["white"], bold=True)
        cell.fill = fill(C["warning"])
        cell.alignment = align("center")
        cell.border = thin
    ws.row_dimensions[cur].height = 20
    row += 1

    for idx, sale in enumerate(sales, 1):
        status = sale.get("payment_status", "")
        status_label = "✅ To'liq" if status == "PAID" else "🔴 Qarz" if status == "DEBT" else "⚠️ Qisman"
        items_count = len(sale.get("items", []))
        dr = ws.append([
            idx,
            fmt_date(sale.get("created_at", "")),
            sale.get("sale_number", ""),
            f"{items_count} ta tovar",
            float(sale.get("total_amount", 0)),
            float(sale.get("paid_amount", 0)),
            float(sale.get("debt_amount", 0)),
            status_label,
            "", "", ""
        ])
        cur = ws._current_row
        for col in range(1, 9):
            cell = ws.cell(cur, col)
            cell.border = thin
            cell.alignment = align("center" if col != 3 else "left")
            if col in (5, 6, 7):
                cell.number_format = "#,##0"
            bg = C["lightGreen"] if status == "PAID" else C["lightRed"] if status == "DEBT" else C["yellow"]
            cell.fill = fill(bg)
        row += 1

    row += 1

    # ── SECTION 4: Debt history - UZS ──
    uzs_records = [r for r in debt_history if r.get("currency", "UZS") != "USD"]
    usd_records = [r for r in debt_history if r.get("currency") == "USD"]

    def render_debt_section(records, title, title_color, header_color, is_usd):
        nonlocal row
        ws.merge_cells(f"A{row}:K{row}")
        c = ws.cell(row, 1)
        c.value = title
        c.font = font(C["white"], bold=True, size=12)
        c.fill = fill(title_color)
        c.alignment = align("center")
        ws.row_dimensions[row].height = 22
        row += 1

        currency_label = "$" if is_usd else "so'm"
        headers = ["№", "Sana", "Vaqt", "Turi", f"Miqdor ({currency_label})",
                   f"Qarz oldin ({currency_label})", f"Qarz keyin ({currency_label})",
                   "O'zgarish", "", "Izoh", ""]
        hr = ws.append(headers)
        cur = ws._current_row
        for col in [1,2,3,4,5,6,7,8,10]:
            cell = ws.cell(cur, col)
            cell.font = font(C["white"], bold=True)
            cell.fill = fill(header_color)
            cell.alignment = align("center")
            cell.border = thin
        ws.row_dimensions[cur].height = 20
        row += 1

        if not records:
            ws.append(["", "", "", "Ma'lumot yo'q", "", "", "", "", "", "", ""])
            ws.cell(ws._current_row, 4).font = font(C["gray"], italic=True)
            row += 1
            return

        total_debt = 0
        total_payment = 0
        for idx, rec in enumerate(records, 1):
            is_payment = rec.get("transaction_type") in ("PAYMENT", "DEBT_PAYMENT")
            amt = abs(float(rec.get("amount", 0)))
            bal_before = float(rec.get("balance_before", 0))
            bal_after = float(rec.get("balance_after", 0))
            change = bal_after - bal_before

            if is_payment:
                total_payment += amt
            else:
                total_debt += amt

            ref = rec.get("reference_type", "")
            ttype = rec.get("transaction_type", "")
            if is_payment:
                type_label = "💰 To'lov"
            elif ttype == "DEBT_INCREASE":
                type_label = "📈 Qarz qo'shildi"
            elif ref in ("adjustment", "adjustment_usd"):
                type_label = "📝 Boshlang'ich qarz"
            else:
                type_label = "📦 Xarid"

            if is_usd:
                amt_str = fmt_usd(amt)
                before_str = fmt_usd(bal_before)
                after_str = fmt_usd(bal_after)
                change_str = (f"+{fmt_usd(abs(change))}" if change >= 0 else f"-{fmt_usd(abs(change))}")
            else:
                amt_str = amt
                before_str = bal_before
                after_str = bal_after
                change_str = change

            ws.append([
                idx,
                fmt_date(rec.get("created_at", "")),
                fmt_time(rec.get("created_at", "")),
                type_label,
                amt_str,
                before_str,
                after_str,
                change_str,
                "",
                rec.get("description", "") or "-",
                ""
            ])
            cur = ws._current_row
            for col in [1,2,3,4,5,6,7,8,10]:
                cell = ws.cell(cur, col)
                cell.border = thin
                cell.alignment = align("right" if col in (5,6,7,8) else "left")
                cell.fill = fill(C["lightGreen"] if is_payment else C["lightRed"])
                if not is_usd and col in (5, 6, 7, 8):
                    cell.number_format = "#,##0"
            # Color change cell
            ws.cell(cur, 8).font = font(C["success"] if change < 0 else C["danger"], bold=True)
            row += 1

        # Totals
        if is_usd:
            total_str = f"Qarz: {fmt_usd(total_debt)}  |  To'lov: {fmt_usd(total_payment)}"
        else:
            total_str = f"Qarz: {fmt_money(total_debt)} so'm  |  To'lov: {fmt_money(total_payment)} so'm"

        ws.append(["", "", "", "JAMI:", total_str, "", "", "", "", "", ""])
        cur = ws._current_row
        for col in (4, 5):
            cell = ws.cell(cur, col)
            cell.font = font(C["blue_dark"] if is_usd else C["danger"], bold=True)
            cell.fill = fill(C["yellow"])
            cell.border = thin
        row += 2

    render_debt_section(
        uzs_records,
        f"💰 SO'M OPERATSIYALARI ({len(uzs_records)} ta yozuv)",
        "BF0000", "C00000", False
    )
    render_debt_section(
        usd_records,
        f"💵 DOLLAR OPERATSIYALARI ({len(usd_records)} ta yozuv)",
        "1E40AF", "2E75B6", True
    )

    # ── Footer ──
    ws.merge_cells(f"A{row}:K{row}")
    fc = ws.cell(row, 1)
    fc.value = "© Vegas System | Hisobot avtomatik tarzda yaratildi"
    fc.font = font(C["gray"], italic=True, size=10)
    fc.alignment = align("center")

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 5
    ws.column_dimensions["J"].width = 45
    ws.column_dimensions["K"].width = 5

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
