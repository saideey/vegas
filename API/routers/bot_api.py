"""
Internal Bot API - endpoints for Telegram Bot service.
No JWT authentication required (internal network only).
These endpoints are called by the Telegram bot to fetch/update customer data.
"""

from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import desc, text as _text

from database import get_db
from database.models import Customer, CustomerDebt, Sale, SaleItem, Payment

router = APIRouter()


@router.get("/customer/by-phone/{phone}")
async def get_customer_by_phone(
        phone: str,
        db: Session = Depends(get_db)
):
    """Find customer by phone number."""
    # Normalize phone: remove spaces, dashes
    clean_phone = phone.strip().replace(" ", "").replace("-", "")

    # Try exact match first
    customer = db.query(Customer).filter(
        Customer.is_deleted == False,
        Customer.phone == clean_phone
    ).first()

    # Try with/without +998 prefix
    if not customer:
        if clean_phone.startswith("+998"):
            alt_phone = clean_phone[4:]  # Remove +998
        elif clean_phone.startswith("998"):
            alt_phone = clean_phone[3:]  # Remove 998
        else:
            alt_phone = "+998" + clean_phone

        customer = db.query(Customer).filter(
            Customer.is_deleted == False,
            (Customer.phone == alt_phone) |
            (Customer.phone == clean_phone) |
            (Customer.phone.contains(clean_phone[-9:]))
        ).first()

    if not customer:
        return {"success": False, "error": "Mijoz topilmadi"}

    return {
        "success": True,
        "data": {
            "id": customer.id,
            "name": customer.name,
            "phone": customer.phone,
            "company_name": customer.company_name,
            "customer_type": customer.customer_type.name if customer.customer_type else "REGULAR",
            "current_debt": float(customer.current_debt or 0),
            "current_debt_usd": float(db.execute(
                _text("SELECT COALESCE(current_debt_usd,0) FROM customers WHERE id=:id"),
                {"id": customer.id}
            ).scalar() or 0),
            "advance_balance": float(customer.advance_balance or 0),
            "credit_limit": float(customer.credit_limit or 0),
            "total_purchases": float(customer.total_purchases or 0),
            "total_purchases_count": customer.total_purchases_count or 0,
            "last_purchase_date": customer.last_purchase_date.isoformat() if customer.last_purchase_date else None,
            "personal_discount_percent": float(customer.personal_discount_percent or 0),
            "telegram_id": customer.telegram_id,
            "address": customer.address,
        }
    }


@router.get("/customer/by-telegram/{telegram_id}")
async def get_customer_by_telegram_id(
        telegram_id: str,
        db: Session = Depends(get_db)
):
    """Find customer by Telegram ID."""
    customer = db.query(Customer).filter(
        Customer.is_deleted == False,
        Customer.telegram_id == telegram_id
    ).first()

    if not customer:
        return {"success": False, "error": "Mijoz topilmadi"}

    return {
        "success": True,
        "data": {
            "id": customer.id,
            "name": customer.name,
            "phone": customer.phone,
            "company_name": customer.company_name,
            "customer_type": customer.customer_type.name if customer.customer_type else "REGULAR",
            "current_debt": float(customer.current_debt or 0),
            "current_debt_usd": float(db.execute(
                _text("SELECT COALESCE(current_debt_usd,0) FROM customers WHERE id=:id"),
                {"id": customer.id}
            ).scalar() or 0),
            "advance_balance": float(customer.advance_balance or 0),
            "credit_limit": float(customer.credit_limit or 0),
            "total_purchases": float(customer.total_purchases or 0),
            "total_purchases_count": customer.total_purchases_count or 0,
            "last_purchase_date": customer.last_purchase_date.isoformat() if customer.last_purchase_date else None,
            "personal_discount_percent": float(customer.personal_discount_percent or 0),
            "telegram_id": customer.telegram_id,
            "address": customer.address,
        }
    }


@router.post("/customer/link-telegram")
async def link_telegram_id(
        data: dict,
        db: Session = Depends(get_db)
):
    """Link Telegram ID to customer account."""
    phone = data.get("phone", "").strip().replace(" ", "").replace("-", "")
    telegram_id = str(data.get("telegram_id", ""))

    if not phone or not telegram_id:
        return {"success": False, "error": "phone va telegram_id majburiy"}

    # Check if telegram_id already linked to another customer
    existing = db.query(Customer).filter(
        Customer.telegram_id == telegram_id,
        Customer.is_deleted == False
    ).first()

    if existing:
        return {
            "success": True,
            "data": {
                "id": existing.id,
                "name": existing.name,
                "phone": existing.phone,
                "already_linked": True
            }
        }

    # Find customer by phone
    clean_phone = phone
    customer = db.query(Customer).filter(
        Customer.is_deleted == False,
        Customer.phone == clean_phone
    ).first()

    # Try alternative phone formats
    if not customer:
        last_9 = clean_phone[-9:] if len(clean_phone) >= 9 else clean_phone
        customer = db.query(Customer).filter(
            Customer.is_deleted == False,
            Customer.phone.contains(last_9)
        ).first()

    if not customer:
        return {"success": False, "error": "Bu telefon raqamli mijoz topilmadi"}

    # Link telegram_id
    customer.telegram_id = telegram_id
    db.commit()

    return {
        "success": True,
        "data": {
            "id": customer.id,
            "name": customer.name,
            "phone": customer.phone,
            "already_linked": False
        }
    }


@router.get("/customer/{customer_id}/info")
async def get_customer_full_info(
        customer_id: int,
        db: Session = Depends(get_db)
):
    """Get full customer info including debt summary."""
    customer = db.query(Customer).filter(
        Customer.id == customer_id,
        Customer.is_deleted == False
    ).first()

    if not customer:
        return {"success": False, "error": "Mijoz topilmadi"}

    return {
        "success": True,
        "data": {
            "id": customer.id,
            "name": customer.name,
            "phone": customer.phone,
            "company_name": customer.company_name,
            "customer_type": customer.customer_type.name if customer.customer_type else "REGULAR",
            "current_debt": float(customer.current_debt or 0),
            "current_debt_usd": float(db.execute(
                _text("SELECT COALESCE(current_debt_usd,0) FROM customers WHERE id=:id"),
                {"id": customer.id}
            ).scalar() or 0),
            "advance_balance": float(customer.advance_balance or 0),
            "credit_limit": float(customer.credit_limit or 0),
            "total_purchases": float(customer.total_purchases or 0),
            "total_purchases_count": customer.total_purchases_count or 0,
            "last_purchase_date": customer.last_purchase_date.isoformat() if customer.last_purchase_date else None,
            "personal_discount_percent": float(customer.personal_discount_percent or 0),
            "address": customer.address,
        }
    }


@router.get("/customer/{customer_id}/purchases")
async def get_customer_purchases(
        customer_id: int,
        page: int = Query(1, ge=1),
        per_page: int = Query(10, ge=1, le=50),
        db: Session = Depends(get_db)
):
    """Get customer purchase history with items."""
    customer = db.query(Customer).filter(
        Customer.id == customer_id,
        Customer.is_deleted == False
    ).first()

    if not customer:
        return {"success": False, "error": "Mijoz topilmadi"}

    # Get sales
    query = db.query(Sale).filter(
        Sale.customer_id == customer_id,
        Sale.is_cancelled == False
    ).order_by(desc(Sale.created_at))

    total = query.count()
    offset = (page - 1) * per_page
    sales = query.offset(offset).limit(per_page).all()

    sales_data = []
    for sale in sales:
        # Get items for this sale
        items = db.query(SaleItem).filter(
            SaleItem.sale_id == sale.id
        ).all()

        items_data = []
        for item in items:
            product_name = item.product.name if item.product else "Noma'lum"
            uom_symbol = item.uom.symbol if item.uom else ""
            items_data.append({
                "product_name": product_name,
                "quantity": float(item.quantity),
                "uom": uom_symbol,
                "unit_price": float(item.unit_price or 0),
                "total_price": float(item.total_price or 0),
            })

        sales_data.append({
            "id": sale.id,
            "sale_number": sale.sale_number,
            "sale_date": sale.sale_date.isoformat() if sale.sale_date else None,
            "created_at": sale.created_at.isoformat() if sale.created_at else None,
            "total_amount": float(sale.total_amount or 0),
            "paid_amount": float(sale.paid_amount or 0),
            "debt_amount": float(sale.debt_amount or 0),
            "discount_amount": float(sale.discount_amount or 0),
            "payment_status": sale.payment_status.name if sale.payment_status else "UNKNOWN",
            "items": items_data,
        })

    return {
        "success": True,
        "data": sales_data,
        "total": total,
        "page": page,
        "per_page": per_page
    }


@router.get("/customer/{customer_id}/debt-details")
async def get_customer_debt_details(
        customer_id: int,
        db: Session = Depends(get_db)
):
    """Get detailed debt info: unpaid sales + payment history."""
    customer = db.query(Customer).filter(
        Customer.id == customer_id,
        Customer.is_deleted == False
    ).first()

    if not customer:
        return {"success": False, "error": "Mijoz topilmadi"}

    # Get unpaid sales (with debt)
    unpaid_sales = db.query(Sale).filter(
        Sale.customer_id == customer_id,
        Sale.is_cancelled == False,
        Sale.debt_amount > 0
    ).order_by(desc(Sale.created_at)).limit(20).all()

    unpaid_data = []
    for sale in unpaid_sales:
        items = db.query(SaleItem).filter(SaleItem.sale_id == sale.id).all()
        items_text = ", ".join([
            f"{item.product.name} ({float(item.quantity)} {item.uom.symbol if item.uom else ''})"
            for item in items[:3]
        ])
        if len(items) > 3:
            items_text += f" +{len(items) - 3} ta"

        unpaid_data.append({
            "sale_number": sale.sale_number,
            "sale_date": sale.sale_date.isoformat() if sale.sale_date else None,
            "total_amount": float(sale.total_amount or 0),
            "paid_amount": float(sale.paid_amount or 0),
            "debt_amount": float(sale.debt_amount or 0),
            "items_summary": items_text,
        })

    # Get recent payment history
    recent_payments = db.query(CustomerDebt).filter(
        CustomerDebt.customer_id == customer_id,
        CustomerDebt.transaction_type.in_(['PAYMENT', 'payment', 'DEBT_PAYMENT'])
    ).order_by(desc(CustomerDebt.created_at)).limit(10).all()

    payments_data = []
    for p in recent_payments:
        payments_data.append({
            "date": p.created_at.isoformat() if p.created_at else None,
            "amount": float(abs(p.amount)),
            "description": p.description,
            "balance_after": float(p.balance_after or 0),
        })

    # Fresh USD debt
    from sqlalchemy import text as _text2
    usd_debt2 = float(db.execute(
        _text2("SELECT COALESCE(current_debt_usd,0) FROM customers WHERE id=:id"),
        {"id": customer_id}
    ).scalar() or 0)

    return {
        "success": True,
        "data": {
            "current_debt": float(customer.current_debt or 0),
            "current_debt_usd": usd_debt2,
            "advance_balance": float(customer.advance_balance or 0),
            "credit_limit": float(customer.credit_limit or 0),
            "unpaid_sales": unpaid_data,
            "recent_payments": payments_data,
        }
    }

@router.get("/customer/{customer_id}/full-report")
async def get_customer_full_report(
    customer_id: int,
    db: Session = Depends(get_db)
):
    """Get all data needed for customer Excel report via bot."""
    from sqlalchemy import text as _text

    customer = db.query(Customer).filter(
        Customer.id == customer_id,
        Customer.is_deleted == False
    ).first()

    if not customer:
        return {"success": False, "message": "Mijoz topilmadi"}

    # Fresh USD debt
    usd_debt = float(db.execute(
        _text("SELECT COALESCE(current_debt_usd,0) FROM customers WHERE id=:id"),
        {"id": customer_id}
    ).scalar() or 0)

    # Sales with items
    from database.models import Sale, SaleItem
    sales = db.query(Sale).filter(
        Sale.customer_id == customer_id,
        Sale.is_cancelled == False
    ).order_by(Sale.created_at.desc()).limit(200).all()

    sales_data = []
    for sale in sales:
        items = db.query(SaleItem).filter(SaleItem.sale_id == sale.id).all()
        items_data = []
        for item in items:
            items_data.append({
                "product_name": item.product_name or "",
                "quantity": float(item.quantity or 0),
                "uom_symbol": item.uom_symbol or "",
                "unit_price": float(item.unit_price or 0),
                "discount": float(item.discount_amount or 0),
                "total": float(item.total_price or 0),
                "article": item.product_article or "",
            })
        sales_data.append({
            "sale_number": sale.sale_number or "",
            "created_at": sale.created_at.isoformat() if sale.created_at else "",
            "total_amount": float(sale.total_amount or 0),
            "paid_amount": float(sale.paid_amount or 0),
            "debt_amount": float(sale.debt_amount or 0),
            "payment_status": sale.payment_status or "",
            "items": items_data,
        })

    # Debt history
    from database.models import CustomerDebt
    debt_records = db.query(CustomerDebt).filter(
        CustomerDebt.customer_id == customer_id
    ).order_by(CustomerDebt.created_at.desc()).limit(200).all()

    debt_data = []
    for r in debt_records:
        debt_data.append({
            "transaction_type": r.transaction_type or "",
            "currency": getattr(r, "currency", "UZS") or "UZS",
            "amount": float(r.amount or 0),
            "balance_before": float(r.balance_before or 0),
            "balance_after": float(r.balance_after or 0),
            "reference_type": r.reference_type or "",
            "description": r.description or "",
            "created_at": r.created_at.isoformat() if r.created_at else "",
        })

    return {
        "success": True,
        "data": {
            "customer": {
                "id": customer.id,
                "name": customer.name,
                "phone": customer.phone,
                "company_name": customer.company_name or "",
                "address": customer.address or "",
                "customer_type": customer.customer_type.name if customer.customer_type else "REGULAR",
                "current_debt": float(customer.current_debt or 0),
                "current_debt_usd": usd_debt,
                "advance_balance": float(customer.advance_balance or 0),
                "total_purchases": float(customer.total_purchases or 0),
                "total_purchases_count": customer.total_purchases_count or 0,
            },
            "sales": sales_data,
            "debt_history": debt_data,
        }
    }


@router.get("/customer/{customer_id}/excel-report")
async def download_customer_excel_report(
    customer_id: int,
    db: Session = Depends(get_db)
):
    """Generate and return customer Excel report as file download."""
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from sqlalchemy import text as _t

    customer = db.query(Customer).filter(
        Customer.id == customer_id,
        Customer.is_deleted == False
    ).first()
    if not customer:
        from fastapi import HTTPException
        raise HTTPException(404, "Mijoz topilmadi")

    # Fresh USD debt
    usd_debt = float(db.execute(
        _t("SELECT COALESCE(current_debt_usd,0) FROM customers WHERE id=:id"),
        {"id": customer_id}
    ).scalar() or 0)

    # Sales
    sales = db.query(Sale).filter(
        Sale.customer_id == customer_id,
        Sale.is_cancelled == False
    ).order_by(Sale.created_at.desc()).limit(200).all()

    # Debt history
    debt_records = db.query(CustomerDebt).filter(
        CustomerDebt.customer_id == customer_id
    ).order_by(CustomerDebt.created_at.desc()).limit(200).all()

    # ── Build Excel ──
    def fmt(n):
        try: return f"{float(n):,.0f}".replace(",", " ")
        except: return "0"

    def fmt_usd(n):
        try: return f"${float(n):,.2f}"
        except: return "$0.00"

    def fdate(iso):
        if not iso: return ""
        try:
            from datetime import datetime
            return datetime.fromisoformat(str(iso)).strftime("%d.%m.%Y")
        except: return str(iso)[:10]

    def ftime(iso):
        if not iso: return ""
        try:
            from datetime import datetime
            return datetime.fromisoformat(str(iso)).strftime("%H:%M")
        except: return ""

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    def fl(c): return PatternFill("solid", fgColor="FF"+c)
    def fn(c="000000", b=False, s=11): return Font(bold=b, size=s, color="FF"+c)
    def al(h="left", v="center"): return Alignment(horizontal=h, vertical=v)

    C = {"hdr":"1F4E79","sub":"2E75B6","ok":"375623","err":"C00000",
         "warn":"FF8C00","wh":"FFFFFF","lb":"DEEAF1","lr":"FCE4D6",
         "lg":"E2EFDA","gr":"808080","yl":"FFF2CC","bl":"1E40AF","bll":"DBEAFE"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Mijoz hisoboti"
    row = 1

    # Header
    ws.merge_cells(f"A{row}:K{row}")
    c = ws.cell(row,1, value="🏢 Vegas - MIJOZ HISOBOTI")
    c.font = fn(C["wh"], True, 16); c.fill = fl(C["hdr"]); c.alignment = al("center")
    ws.row_dimensions[row].height = 32
    row += 2

    # Customer info
    debt_uzs = float(customer.current_debt or 0)
    total_sales_count = len(sales)
    total_amount = sum(float(s.total_amount or 0) for s in sales)
    total_paid = sum(float(s.paid_amount or 0) for s in sales)
    total_debt_sales = sum(float(s.debt_amount or 0) for s in sales)
    advance = float(customer.advance_balance or 0)

    ws.merge_cells(f"A{row}:E{row}")
    h1 = ws.cell(row,1, value="👤 MIJOZ MA'LUMOTLARI")
    h1.font=fn(C["wh"],True,12); h1.fill=fl(C["sub"]); h1.alignment=al("center")
    ws.merge_cells(f"F{row}:K{row}")
    h2 = ws.cell(row,6, value="📊 MOLIYAVIY STATISTIKA")
    h2.font=fn(C["wh"],True,12); h2.fill=fl(C["ok"]); h2.alignment=al("center")
    row += 1

    info = [
        ["Mijoz:", customer.name, "", "", "", "Jami xaridlar:", f"{total_sales_count} ta"],
        ["Telefon:", customer.phone, "", "", "", "Jami summa:", fmt(total_amount)+" so'm"],
        ["Kompaniya:", customer.company_name or "-", "", "", "", "To'langan:", fmt(total_paid)+" so'm"],
        ["Manzil:", customer.address or "-", "", "", "", "Qarz (sotuvlardan):", fmt(total_debt_sales)+" so'm"],
        ["", "", "", "", "", "Joriy qarz (so'm):", (fmt(debt_uzs)+" so'm") if debt_uzs>0 else "0 so'm"],
        ["", "", "", "", "", "Joriy qarz ($):", fmt_usd(usd_debt)],
        ["", "", "", "", "", "Avans balansi:", fmt(advance)+" so'm"],
    ]
    debt_colors = {4: (C["err"] if debt_uzs>0 else C["ok"], C["lr"] if debt_uzs>0 else C["lg"]),
                   5: (C["bl"], C["bll"] if usd_debt>0 else C["lg"]),
                   6: (C["ok"], C["lg"])}
    for i, ir in enumerate(info):
        ws.append(ir)
        cur = ws._current_row
        ws.cell(cur,1).font=fn(C["sub"],True); ws.cell(cur,1).fill=fl(C["lb"])
        ws.cell(cur,2).fill=fl(C["wh"])
        if i in debt_colors:
            fc, bg = debt_colors[i]
            ws.cell(cur,6).font=fn(fc,True); ws.cell(cur,6).fill=fl(C["lb"])
            ws.cell(cur,7).font=fn(fc,True); ws.cell(cur,7).fill=fl(bg)
        else:
            ws.cell(cur,6).font=fn(C["ok"],True); ws.cell(cur,6).fill=fl(C["lg"])
        row += 1
    row += 1

    # Sales section
    ws.merge_cells(f"A{row}:K{row}")
    st = ws.cell(row,1, value=f"🛒 SOTUVLAR TARIXI ({total_sales_count} ta)")
    st.font=fn(C["wh"],True,12); st.fill=fl(C["warn"]); st.alignment=al("center")
    ws.row_dimensions[row].height=22; row+=1

    for h,v in zip(["№","Sana","Chek №","Jami","To'langan","Qarz","Holat","","","",""],range(1,12)):
        c=ws.cell(row,v,value=h); c.font=fn(C["wh"],True); c.fill=fl(C["warn"]); c.alignment=al("center"); c.border=thin
    ws.row_dimensions[row].height=20; row+=1

    for idx,sale in enumerate(sales,1):
        status = sale.payment_status.name if sale.payment_status else ""
        sl = "✅ To'liq" if status=="PAID" else "🔴 Qarz" if status=="DEBT" else "⚠️ Qisman"
        bg = C["lg"] if status=="PAID" else C["lr"] if status=="DEBT" else C["yl"]
        vals = [idx, fdate(sale.created_at), sale.sale_number or "",
                float(sale.total_amount or 0), float(sale.paid_amount or 0),
                float(sale.debt_amount or 0), sl, "", "", "", ""]
        ws.append(vals)
        cur=ws._current_row
        for col in range(1,8):
            cell=ws.cell(cur,col); cell.border=thin; cell.fill=fl(bg)
            cell.alignment=al("right" if col>=4 else "center")
            if col in(4,5,6): cell.number_format="#,##0"
        row+=1
    row+=1

    # Debt history — split UZS / USD
    uzs_recs = [r for r in debt_records if not getattr(r,"currency",None) or r.currency=="UZS"]
    usd_recs = [r for r in debt_records if getattr(r,"currency",None)=="USD"]

    def debt_section(recs, title, tc, hc, is_usd):
        nonlocal row
        ws.merge_cells(f"A{row}:K{row}")
        c=ws.cell(row,1,value=title); c.font=fn(C["wh"],True,12)
        c.fill=fl(tc); c.alignment=al("center"); ws.row_dimensions[row].height=22; row+=1
        lbl="$" if is_usd else "so'm"
        hdrs=["№","Sana","Vaqt","Turi",f"Miqdor({lbl})",f"Qarz oldin({lbl})",f"Qarz keyin({lbl})","O'zgarish","","Izoh",""]
        ws.append(hdrs)
        cur=ws._current_row
        for col in [1,2,3,4,5,6,7,8,10]:
            c=ws.cell(cur,col); c.font=fn(C["wh"],True); c.fill=fl(hc)
            c.alignment=al("center"); c.border=thin
        ws.row_dimensions[cur].height=20; row+=1
        if not recs:
            ws.append(["","","","Ma'lumot yo'q","","","","","","",""])
            ws.cell(ws._current_row,4).font=fn(C["gr"],False); row+=1; return
        td=tp=0
        for idx,r in enumerate(recs,1):
            is_pay=r.transaction_type in("PAYMENT","DEBT_PAYMENT")
            amt=abs(float(r.amount or 0))
            bb=float(r.balance_before or 0); ba=float(r.balance_after or 0); ch=ba-bb
            if is_pay: tp+=amt
            else: td+=amt
            rt=r.reference_type or ""; tt=r.transaction_type or ""
            if is_pay: tl="💰 To'lov"
            elif tt=="DEBT_INCREASE": tl="📈 Qarz qo'shildi"
            elif rt in("adjustment","adjustment_usd"): tl="📝 Boshlang'ich"
            else: tl="📦 Xarid"
            if is_usd:
                row_vals=[idx,fdate(r.created_at),ftime(r.created_at),tl,
                          fmt_usd(amt),fmt_usd(bb),fmt_usd(ba),
                          ("+"+fmt_usd(abs(ch))) if ch>=0 else ("-"+fmt_usd(abs(ch))),
                          "",r.description or "-",""]
            else:
                row_vals=[idx,fdate(r.created_at),ftime(r.created_at),tl,
                          amt,bb,ba,ch,"",r.description or "-",""]
            ws.append(row_vals)
            cur=ws._current_row
            bg=C["lg"] if is_pay else C["lr"]
            for col in [1,2,3,4,5,6,7,8,10]:
                cell=ws.cell(cur,col); cell.border=thin
                cell.fill=fl(bg); cell.alignment=al("right" if col in(5,6,7,8) else "left")
                if not is_usd and col in(5,6,7,8): cell.number_format="#,##0"
            ws.cell(cur,8).font=fn(C["ok"] if ch<0 else C["err"],True); row+=1
        # totals
        if is_usd: ts=f"Qarz:{fmt_usd(td)}  To'lov:{fmt_usd(tp)}"
        else: ts=f"Qarz:{fmt(td)} so'm  To'lov:{fmt(tp)} so'm"
        ws.append(["","","","JAMI:",ts,"","","","","",""])
        cur=ws._current_row
        for col in(4,5):
            c=ws.cell(cur,col); c.font=fn(C["bl"] if is_usd else C["err"],True)
            c.fill=fl(C["yl"]); c.border=thin
        row+=2

    debt_section(uzs_recs, f"💰 SO'M OPERATSIYALARI ({len(uzs_recs)} ta)", "BF0000","C00000", False)
    debt_section(usd_recs, f"💵 DOLLAR OPERATSIYALARI ({len(usd_recs)} ta)", "1E40AF","2E75B6", True)

    # Footer
    ws.merge_cells(f"A{row}:K{row}")
    fc=ws.cell(row,1,value="© Vegas System")
    fc.font=fn(C["gr"],False,10); fc.alignment=al("center")

    # Column widths
    for col,w in zip("ABCDEFGHIJK",[6,14,10,28,18,18,18,18,5,45,5]):
        ws.column_dimensions[col].width=w

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    from datetime import datetime
    fname=f"{customer.name.replace(' ','_')}_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )
