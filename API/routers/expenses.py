"""
Expenses router - CRUD operations for business expenses.
Tracks all expenses: rent, salaries, utilities, etc.
"""

from typing import Optional
from decimal import Decimal
from datetime import date, datetime
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
from sqlalchemy import func
from io import BytesIO

from database import get_db
from database.models import User, PermissionType, Expense, ExpenseCategory
from database.base import get_tashkent_now
from core.dependencies import get_current_active_user, PermissionChecker
from pydantic import BaseModel as PydanticBase, field_validator

router = APIRouter()


# ==================== SCHEMAS ====================

class ExpenseCategoryCreate(PydanticBase):
    name: str
    description: Optional[str] = None

class ExpenseCreate(PydanticBase):
    expense_date: date
    category_id: Optional[int] = None
    amount: Decimal
    payment_type: str = "cash"
    description: str

    @field_validator("category_id", mode="before")
    @classmethod
    def clean_category_id(cls, v):
        if v == "" or v == "undefined" or v is None:
            return None
        return int(v) if v else None

class ExpenseUpdate(PydanticBase):
    expense_date: Optional[date] = None
    category_id: Optional[int] = None
    amount: Optional[Decimal] = None
    payment_type: Optional[str] = None
    description: Optional[str] = None


# ==================== CATEGORIES ====================

@router.get("/categories", summary="Chiqim kategoriyalari")
async def get_expense_categories(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get all expense categories."""
    categories = db.query(ExpenseCategory).filter(
        ExpenseCategory.is_active == True
    ).order_by(ExpenseCategory.name).all()

    return {
        "success": True,
        "data": [{
            "id": c.id,
            "name": c.name,
            "description": c.description
        } for c in categories]
    }


@router.post("/categories", summary="Kategoriya qo'shish")
async def create_expense_category(
    data: ExpenseCategoryCreate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Create a new expense category."""
    existing = db.query(ExpenseCategory).filter(
        ExpenseCategory.name == data.name
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Bu kategoriya allaqachon mavjud")

    category = ExpenseCategory(
        name=data.name,
        description=data.description,
        is_active=True
    )
    db.add(category)
    db.commit()
    db.refresh(category)

    return {
        "success": True,
        "data": {"id": category.id, "name": category.name},
        "message": "Kategoriya yaratildi"
    }


@router.delete("/categories/{category_id}", summary="Kategoriya o'chirish")
async def delete_expense_category(
    category_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Delete (deactivate) expense category."""
    category = db.query(ExpenseCategory).filter(ExpenseCategory.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Kategoriya topilmadi")

    category.is_active = False
    db.commit()
    return {"success": True, "message": "Kategoriya o'chirildi"}


# ==================== EXPENSES CRUD ====================

@router.get("", summary="Chiqimlar ro'yxati")
async def get_expenses(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    category_id: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get paginated expenses list with filters."""
    query = db.query(Expense)

    if start_date:
        query = query.filter(Expense.expense_date >= start_date)
    if end_date:
        query = query.filter(Expense.expense_date <= end_date)
    if category_id:
        query = query.filter(Expense.category_id == category_id)

    total = query.count()

    # Summary for filtered period
    summary = query.with_entities(
        func.coalesce(func.sum(Expense.amount), 0)
    ).scalar()

    expenses = query.order_by(Expense.expense_date.desc(), Expense.created_at.desc())\
        .offset((page - 1) * per_page).limit(per_page).all()

    data = [{
        "id": e.id,
        "expense_date": e.expense_date.isoformat(),
        "category_id": e.category_id,
        "category_name": e.category.name if e.category else "Boshqa",
        "amount": float(e.amount or 0),
        "payment_type": e.payment_type,
        "description": e.description,
        "created_by_name": f"{e.created_by.first_name} {e.created_by.last_name}" if e.created_by else None,
        "created_at": e.created_at.isoformat() if e.created_at else None
    } for e in expenses]

    return {
        "success": True,
        "data": data,
        "total": total,
        "total_amount": float(summary or 0),
        "page": page,
        "per_page": per_page
    }


@router.post("", summary="Chiqim qo'shish")
async def create_expense(
    data: ExpenseCreate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Create a new expense record."""
    if data.amount <= 0:
        raise HTTPException(status_code=400, detail="Summa 0 dan katta bo'lishi kerak")

    expense = Expense(
        expense_date=data.expense_date,
        category_id=data.category_id,
        amount=data.amount,
        payment_type=data.payment_type,
        description=data.description,
        created_by_id=current_user.id
    )
    db.add(expense)
    db.commit()
    db.refresh(expense)

    return {
        "success": True,
        "data": {
            "id": expense.id,
            "amount": float(expense.amount),
            "description": expense.description
        },
        "message": f"Chiqim yozildi: {float(expense.amount):,.0f} so'm"
    }


@router.put("/{expense_id}", summary="Chiqim tahrirlash")
async def update_expense(
    expense_id: int,
    data: ExpenseUpdate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Update an existing expense."""
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail="Chiqim topilmadi")

    if data.expense_date is not None:
        expense.expense_date = data.expense_date
    if data.category_id is not None:
        expense.category_id = data.category_id
    if data.amount is not None:
        if data.amount <= 0:
            raise HTTPException(status_code=400, detail="Summa 0 dan katta bo'lishi kerak")
        expense.amount = data.amount
    if data.payment_type is not None:
        expense.payment_type = data.payment_type
    if data.description is not None:
        expense.description = data.description

    db.commit()
    return {"success": True, "message": "Chiqim yangilandi"}


@router.delete("/{expense_id}", summary="Chiqim o'chirish")
async def delete_expense(
    expense_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Delete an expense record."""
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail="Chiqim topilmadi")

    db.delete(expense)
    db.commit()
    return {"success": True, "message": "Chiqim o'chirildi"}


# ==================== REPORTS ====================

@router.get("/summary", summary="Chiqimlar hisoboti")
async def get_expenses_summary(
    start_date: date = None,
    end_date: date = None,
    period: str = Query("month", regex="^(day|week|month|year)$"),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Get expenses summary by category for a given period.
    Period: day, week, month, year.
    """
    from datetime import timedelta

    today = date.today()

    if not start_date or not end_date:
        if period == "day":
            start_date = today
            end_date = today
        elif period == "week":
            start_date = today - timedelta(days=today.weekday())
            end_date = today
        elif period == "month":
            start_date = today.replace(day=1)
            end_date = today
        elif period == "year":
            start_date = today.replace(month=1, day=1)
            end_date = today

    # Total expenses
    total_expenses = db.query(
        func.coalesce(func.sum(Expense.amount), 0)
    ).filter(
        Expense.expense_date >= start_date,
        Expense.expense_date <= end_date
    ).scalar() or 0

    # By category
    category_breakdown = db.query(
        ExpenseCategory.id,
        ExpenseCategory.name,
        func.coalesce(func.sum(Expense.amount), 0).label('total'),
        func.count(Expense.id).label('count')
    ).outerjoin(Expense, (Expense.category_id == ExpenseCategory.id) & 
                (Expense.expense_date >= start_date) & 
                (Expense.expense_date <= end_date))\
     .filter(ExpenseCategory.is_active == True)\
     .group_by(ExpenseCategory.id, ExpenseCategory.name)\
     .order_by(func.sum(Expense.amount).desc().nullslast()).all()

    # Uncategorized
    uncategorized = db.query(
        func.coalesce(func.sum(Expense.amount), 0).label('total'),
        func.count(Expense.id).label('count')
    ).filter(
        Expense.expense_date >= start_date,
        Expense.expense_date <= end_date,
        Expense.category_id == None
    ).first()

    # By payment type
    payment_breakdown = db.query(
        Expense.payment_type,
        func.coalesce(func.sum(Expense.amount), 0).label('total')
    ).filter(
        Expense.expense_date >= start_date,
        Expense.expense_date <= end_date
    ).group_by(Expense.payment_type).all()

    # Gross profit for the period (from sales)
    from database.models import Sale, SaleItem, Product as ProductModel
    from sqlalchemy import case as sa_case
    _eff_cost = sa_case((SaleItem.unit_cost > 0, SaleItem.unit_cost), else_=ProductModel.cost_price)
    profit_data = db.query(
        func.coalesce(func.sum(SaleItem.total_price), 0).label('revenue'),
        func.coalesce(func.sum(_eff_cost * SaleItem.base_quantity), 0).label('cost')
    ).join(Sale, Sale.id == SaleItem.sale_id)\
     .join(ProductModel, ProductModel.id == SaleItem.product_id)\
     .filter(
        Sale.sale_date >= start_date,
        Sale.sale_date <= end_date,
        Sale.is_cancelled == False
    ).first()

    gross_profit = float(profit_data.revenue or 0) - float(profit_data.cost or 0) if profit_data else 0
    net_profit = gross_profit - float(total_expenses)

    categories = [{
        "id": c.id,
        "name": c.name,
        "total": float(c.total or 0),
        "count": c.count or 0
    } for c in category_breakdown if float(c.total or 0) > 0]

    if uncategorized and float(uncategorized.total or 0) > 0:
        categories.append({
            "id": None,
            "name": "Boshqa",
            "total": float(uncategorized.total or 0),
            "count": uncategorized.count or 0
        })

    payments = {p.payment_type: float(p.total or 0) for p in payment_breakdown}

    return {
        "success": True,
        "data": {
            "total_expenses": float(total_expenses),
            "gross_profit": gross_profit,
            "net_profit": net_profit,
            "categories": categories,
            "payment_breakdown": payments,
            "period": {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "type": period
            }
        }
    }


@router.get("/export", summary="Chiqimlar Excel export")
async def export_expenses(
    start_date: date = None,
    end_date: date = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export expenses to Excel."""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    today = date.today()
    if not start_date:
        start_date = today.replace(day=1)
    if not end_date:
        end_date = today

    expenses = db.query(Expense).filter(
        Expense.expense_date >= start_date,
        Expense.expense_date <= end_date
    ).order_by(Expense.expense_date.desc(), Expense.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Chiqimlar"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = f"CHIQIMLAR HISOBOTI"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"{start_date.strftime('%d.%m.%Y')} — {end_date.strftime('%d.%m.%Y')}"

    headers = ["№", "Sana", "Kategoriya", "Summa", "To'lov turi", "Izoh", "Kim yozdi", "Vaqt"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    payment_labels = {'cash': 'Naqd', 'card': 'Karta', 'transfer': "O'tkazma"}
    total_amount = 0

    for idx, e in enumerate(expenses, 1):
        row = idx + 4
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=e.expense_date.strftime('%d.%m.%Y')).border = border
        ws.cell(row=row, column=3, value=e.category.name if e.category else "Boshqa").border = border
        amount_cell = ws.cell(row=row, column=4, value=float(e.amount or 0))
        amount_cell.border = border
        amount_cell.number_format = '#,##0'
        ws.cell(row=row, column=5, value=payment_labels.get(e.payment_type, e.payment_type)).border = border
        ws.cell(row=row, column=6, value=e.description or "").border = border
        ws.cell(row=row, column=7, value=f"{e.created_by.first_name} {e.created_by.last_name}" if e.created_by else "").border = border
        ws.cell(row=row, column=8, value=e.created_at.strftime('%d.%m.%Y %H:%M') if e.created_at else "").border = border
        total_amount += float(e.amount or 0)

    # Total row
    total_row = len(expenses) + 5
    ws.cell(row=total_row, column=3, value="JAMI:").font = Font(bold=True, size=12)
    total_cell = ws.cell(row=total_row, column=4, value=total_amount)
    total_cell.font = Font(bold=True, size=12, color="DC143C")
    total_cell.number_format = '#,##0'

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"chiqimlar_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
