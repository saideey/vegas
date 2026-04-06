"""
Reports router - Excel/PDF report generation and download.
"""

from datetime import date
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io

from database import get_db
from database.models import User, PermissionType
from core.dependencies import get_current_active_user, PermissionChecker
from services.reports import ExcelReportGenerator, PDFReportGenerator


router = APIRouter()


# ==================== EXCEL REPORTS ====================

@router.get(
    "/excel/sales",
    summary="Sotuvlar hisoboti (Excel)",
    dependencies=[Depends(PermissionChecker([PermissionType.REPORT_SALES]))]
)
async def export_sales_excel(
    start_date: date,
    end_date: date,
    warehouse_id: Optional[int] = None,
    seller_id: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export sales report as Excel file."""
    generator = ExcelReportGenerator(db)
    
    try:
        excel_data = generator.generate_sales_report(
            start_date=start_date,
            end_date=end_date,
            warehouse_id=warehouse_id,
            seller_id=seller_id
        )
        
        filename = f"sotuvlar_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Hisobot yaratishda xato: {str(e)}")


@router.get(
    "/excel/stock",
    summary="Qoldiqlar hisoboti (Excel)",
    dependencies=[Depends(PermissionChecker([PermissionType.REPORT_WAREHOUSE]))]
)
async def export_stock_excel(
    warehouse_id: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export stock/inventory report as Excel file."""
    generator = ExcelReportGenerator(db)
    
    try:
        excel_data = generator.generate_stock_report(warehouse_id=warehouse_id)
        filename = f"qoldiqlar_{date.today().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Hisobot yaratishda xato: {str(e)}")


@router.get(
    "/excel/debtors",
    summary="Qarzdorlar hisoboti (Excel)",
    dependencies=[Depends(PermissionChecker([PermissionType.REPORT_FINANCE]))]
)
async def export_debtors_excel(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export debtors report as Excel file."""
    generator = ExcelReportGenerator(db)
    
    try:
        excel_data = generator.generate_debtors_report()
        filename = f"qarzdorlar_{date.today().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Hisobot yaratishda xato: {str(e)}")


@router.get(
    "/excel/daily",
    summary="Kunlik hisobot (Excel)",
    dependencies=[Depends(PermissionChecker([PermissionType.REPORT_SALES]))]
)
async def export_daily_excel(
    report_date: Optional[date] = None,
    warehouse_id: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export daily summary report as Excel file."""
    generator = ExcelReportGenerator(db)
    
    if not report_date:
        report_date = date.today()
    
    try:
        excel_data = generator.generate_daily_report(
            report_date=report_date,
            warehouse_id=warehouse_id
        )
        filename = f"kunlik_{report_date.strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Hisobot yaratishda xato: {str(e)}")


@router.get(
    "/excel/price-list",
    summary="Narxlar ro'yxati (Excel)"
)
async def export_price_list_excel(
    category_id: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export product price list as Excel file."""
    generator = ExcelReportGenerator(db)
    
    try:
        excel_data = generator.generate_products_price_list(category_id=category_id)
        filename = f"narxlar_{date.today().strftime('%Y%m%d')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Hisobot yaratishda xato: {str(e)}")


# ==================== PDF REPORTS ====================

@router.get(
    "/pdf/receipt/{sale_id}",
    summary="Chek (PDF)"
)
async def get_receipt_pdf(
    sale_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Generate and download sale receipt as PDF."""
    generator = PDFReportGenerator(db)
    
    try:
        pdf_data = generator.generate_receipt(sale_id)
        
        return StreamingResponse(
            io.BytesIO(pdf_data),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=chek_{sale_id}.pdf"}
        )
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Chek yaratishda xato: {str(e)}")


@router.get(
    "/pdf/sales",
    summary="Sotuvlar hisoboti (PDF)",
    dependencies=[Depends(PermissionChecker([PermissionType.REPORT_SALES]))]
)
async def export_sales_pdf(
    start_date: date,
    end_date: date,
    warehouse_id: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export sales report as PDF file."""
    generator = PDFReportGenerator(db)
    
    try:
        pdf_data = generator.generate_sales_report(
            start_date=start_date,
            end_date=end_date,
            warehouse_id=warehouse_id
        )
        
        filename = f"sotuvlar_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.pdf"
        
        return StreamingResponse(
            io.BytesIO(pdf_data),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Hisobot yaratishda xato: {str(e)}")


@router.get(
    "/pdf/debtors",
    summary="Qarzdorlar hisoboti (PDF)",
    dependencies=[Depends(PermissionChecker([PermissionType.REPORT_FINANCE]))]
)
async def export_debtors_pdf(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export debtors report as PDF file."""
    generator = PDFReportGenerator(db)
    
    try:
        pdf_data = generator.generate_debtors_report()
        filename = f"qarzdorlar_{date.today().strftime('%Y%m%d')}.pdf"
        
        return StreamingResponse(
            io.BytesIO(pdf_data),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Hisobot yaratishda xato: {str(e)}")


@router.get(
    "/pdf/stock",
    summary="Qoldiqlar hisoboti (PDF)",
    dependencies=[Depends(PermissionChecker([PermissionType.REPORT_WAREHOUSE]))]
)
async def export_stock_pdf(
    warehouse_id: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export stock report as PDF file."""
    generator = PDFReportGenerator(db)
    
    try:
        pdf_data = generator.generate_stock_report(warehouse_id=warehouse_id)
        filename = f"qoldiqlar_{date.today().strftime('%Y%m%d')}.pdf"
        
        return StreamingResponse(
            io.BytesIO(pdf_data),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Hisobot yaratishda xato: {str(e)}")


@router.get(
    "/pdf/price-list",
    summary="Narxlar ro'yxati (PDF)"
)
async def export_price_list_pdf(
    category_id: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export product price list as PDF file."""
    # For now, redirect to Excel since PDF is more complex
    raise HTTPException(
        status_code=501, 
        detail="PDF narxlar ro'yxati hali tayyor emas. Excel formatidan foydalaning."
    )


# ==================== JSON DATA REPORTS ====================

@router.get(
    "/profit",
    summary="Foyda hisoboti (JSON)",
    dependencies=[Depends(PermissionChecker([PermissionType.REPORT_PROFIT]))]
)
async def get_profit_report(
    start_date: date,
    end_date: date,
    warehouse_id: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get profit report by product - shows cost, revenue, and profit per product."""
    from sqlalchemy import func, case
    from decimal import Decimal
    from database.models import Sale, SaleItem, Product, Expense, ExpenseCategory, Stock
    
    # Query sale items with product info
    # Agar SaleItem.unit_cost = 0 bo'lsa, product.cost_price dan foydalanamiz
    effective_cost = case(
        (SaleItem.unit_cost > 0, SaleItem.unit_cost),
        else_=Product.cost_price
    )
    
    query = db.query(
        SaleItem.product_id,
        Product.name.label('product_name'),
        Product.article.label('product_article'),
        func.sum(SaleItem.quantity).label('total_quantity'),
        func.sum(effective_cost * SaleItem.base_quantity).label('total_cost'),
        func.sum(SaleItem.total_price).label('total_revenue'),
        func.sum(SaleItem.total_price - (effective_cost * SaleItem.base_quantity)).label('total_profit')
    ).join(Sale, Sale.id == SaleItem.sale_id)\
     .join(Product, Product.id == SaleItem.product_id)\
     .filter(
        Sale.sale_date >= start_date,
        Sale.sale_date <= end_date,
        Sale.is_cancelled == False
    )
    
    if warehouse_id:
        query = query.filter(Sale.warehouse_id == warehouse_id)
    
    results = query.group_by(
        SaleItem.product_id,
        Product.name,
        Product.article
    ).order_by(func.sum(SaleItem.total_price - (effective_cost * SaleItem.base_quantity)).desc()).all()
    
    # Calculate totals
    total_cost = sum(r.total_cost or Decimal("0") for r in results)
    total_revenue = sum(r.total_revenue or Decimal("0") for r in results)
    total_profit = sum(r.total_profit or Decimal("0") for r in results)
    
    data = []
    for r in results:
        profit = r.total_profit or Decimal("0")
        revenue = r.total_revenue or Decimal("1")
        margin = (profit / revenue * 100) if revenue > 0 else Decimal("0")
        
        data.append({
            "product_id": r.product_id,
            "product_name": r.product_name,
            "product_article": r.product_article,
            "total_quantity": float(r.total_quantity or 0),
            "total_cost": float(r.total_cost or 0),
            "total_revenue": float(r.total_revenue or 0),
            "total_profit": float(profit),
            "profit_margin": float(margin)
        })
    
    # ===== CHIQIMLAR =====
    total_expenses_amount = float(db.query(
        func.coalesce(func.sum(Expense.amount), 0)
    ).filter(
        Expense.expense_date >= start_date,
        Expense.expense_date <= end_date
    ).scalar() or 0)

    # Expenses by category
    expenses_by_cat = db.query(
        ExpenseCategory.name,
        func.coalesce(func.sum(Expense.amount), 0).label('total'),
        func.count(Expense.id).label('count')
    ).outerjoin(Expense, (Expense.category_id == ExpenseCategory.id) &
                (Expense.expense_date >= start_date) &
                (Expense.expense_date <= end_date))\
     .filter(ExpenseCategory.is_active == True)\
     .group_by(ExpenseCategory.id, ExpenseCategory.name)\
     .having(func.sum(Expense.amount) > 0).all()

    expenses_list = [{"name": e.name, "total": float(e.total or 0), "count": e.count} for e in expenses_by_cat]

    # Uncategorized
    uncat = float(db.query(func.coalesce(func.sum(Expense.amount), 0)).filter(
        Expense.expense_date >= start_date,
        Expense.expense_date <= end_date,
        Expense.category_id == None
    ).scalar() or 0)
    if uncat > 0:
        expenses_list.append({"name": "Boshqa", "total": uncat, "count": 0})

    net_profit = float(total_profit) - total_expenses_amount

    return {
        "success": True,
        "data": data,
        "summary": {
            "total_cost": float(total_cost),
            "total_revenue": float(total_revenue),
            "total_profit": float(total_profit),
            "profit_margin": float((total_profit / total_revenue * 100) if total_revenue > 0 else 0),
            "total_expenses": total_expenses_amount,
            "net_profit": net_profit,
            "expenses_list": expenses_list,
            "products_count": len(data)
        },
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat()
        }
    }


@router.get(
    "/sales-summary",
    summary="Sotuvlar xulosasi (JSON)",
    dependencies=[Depends(PermissionChecker([PermissionType.REPORT_SALES]))]
)
async def get_sales_summary(
    start_date: date,
    end_date: date,
    warehouse_id: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get sales summary with payment breakdown."""
    from sqlalchemy import func
    from database.models import Sale, Payment, PaymentType
    
    # Base query
    query = db.query(Sale).filter(
        Sale.sale_date >= start_date,
        Sale.sale_date <= end_date,
        Sale.is_cancelled == False
    )
    
    if warehouse_id:
        query = query.filter(Sale.warehouse_id == warehouse_id)
    
    sales = query.all()
    
    # Calculate totals
    total_sales = len(sales)
    total_amount = sum(s.total_amount for s in sales)
    total_discount = sum(s.discount_amount for s in sales)
    total_paid = sum(s.paid_amount for s in sales)
    total_debt = sum(s.debt_amount for s in sales)
    
    # Payment breakdown
    payment_query = db.query(
        Payment.payment_type,
        func.sum(Payment.amount).label('total')
    ).filter(
        Payment.payment_date >= start_date,
        Payment.payment_date <= end_date,
        Payment.is_cancelled == False
    ).group_by(Payment.payment_type).all()
    
    payment_breakdown = {
        pt.value if hasattr(pt, 'value') else str(pt): float(total or 0)
        for pt, total in payment_query
    }
    
    return {
        "success": True,
        "summary": {
            "total_sales": total_sales,
            "total_amount": float(total_amount),
            "total_discount": float(total_discount),
            "total_paid": float(total_paid),
            "total_debt": float(total_debt),
            "average_sale": float(total_amount / total_sales) if total_sales > 0 else 0
        },
        "payment_breakdown": payment_breakdown,
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat()
        }
    }


@router.get(
    "/seller-stats",
    summary="Kassir statistikasi (JSON)"
)
async def get_seller_stats(
    seller_id: int,
    start_date: date,
    end_date: date,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get detailed statistics for a specific seller/cashier."""
    from sqlalchemy import func
    from decimal import Decimal
    from database.models import Sale, SaleItem, Payment, PaymentType, User as UserModel, Customer
    
    # Get seller info
    seller = db.query(UserModel).filter(UserModel.id == seller_id).first()
    if not seller:
        raise HTTPException(status_code=404, detail="Kassir topilmadi")
    
    # Get all sales by this seller
    sales_query = db.query(Sale).filter(
        Sale.seller_id == seller_id,
        Sale.sale_date >= start_date,
        Sale.sale_date <= end_date,
        Sale.is_cancelled == False
    )
    
    sales = sales_query.all()
    
    # Summary stats
    total_sales_count = len(sales)
    total_amount = sum(s.total_amount for s in sales) if sales else Decimal("0")
    total_paid = sum(s.paid_amount for s in sales) if sales else Decimal("0")
    total_debt = sum(s.debt_amount for s in sales) if sales else Decimal("0")
    total_discount = sum(s.discount_amount for s in sales) if sales else Decimal("0")
    
    # Payment breakdown for this seller - directly from Sales table
    # PaymentType enum values are lowercase: cash, card, transfer
    payment_breakdown = {
        "cash": 0.0,
        "card": 0.0,
        "transfer": 0.0
    }
    
    # Calculate from sales directly since Sale table has payment_type
    for sale in sales:
        if sale.paid_amount and float(sale.paid_amount) > 0 and sale.payment_type:
            pt_key = sale.payment_type.value if hasattr(sale.payment_type, 'value') else str(sale.payment_type)
            if pt_key in payment_breakdown:
                payment_breakdown[pt_key] += float(sale.paid_amount)
    
    # Convert to uppercase keys for frontend compatibility
    payment_breakdown = {
        "CASH": payment_breakdown.get("cash", 0.0),
        "CARD": payment_breakdown.get("card", 0.0),
        "TRANSFER": payment_breakdown.get("transfer", 0.0)
    }
    
    # Customers served by this seller
    customers_query = db.query(
        Customer.id,
        Customer.name,
        Customer.phone,
        Customer.company_name,
        Customer.customer_type,
        func.count(Sale.id).label('sales_count'),
        func.sum(Sale.total_amount).label('total_amount'),
        func.sum(Sale.debt_amount).label('total_debt')
    ).join(Sale, Sale.customer_id == Customer.id)\
     .filter(
        Sale.seller_id == seller_id,
        Sale.sale_date >= start_date,
        Sale.sale_date <= end_date,
        Sale.is_cancelled == False
    ).group_by(
        Customer.id,
        Customer.name,
        Customer.phone,
        Customer.company_name,
        Customer.customer_type
    ).order_by(func.sum(Sale.total_amount).desc()).all()
    
    customers = [{
        "id": c.id,
        "name": c.name,
        "phone": c.phone,
        "company_name": c.company_name,
        "customer_type": c.customer_type.name if c.customer_type else "REGULAR",
        "sales_count": c.sales_count,
        "total_amount": float(c.total_amount or 0),
        "total_debt": float(c.total_debt or 0)
    } for c in customers_query]
    
    # Count sales without customer (anonymous)
    anonymous_sales = sum(1 for s in sales if not s.customer_id)
    anonymous_amount = sum(s.total_amount for s in sales if not s.customer_id)
    
    # Recent sales list
    recent_sales = sales_query.order_by(Sale.created_at.desc()).limit(50).all()
    sales_list = [{
        "id": s.id,
        "sale_number": s.sale_number,
        "sale_date": s.sale_date.isoformat(),
        "customer_name": s.customer.name if s.customer else "Oddiy xaridor",
        "customer_id": s.customer_id,
        "total_amount": float(s.total_amount),
        "paid_amount": float(s.paid_amount),
        "debt_amount": float(s.debt_amount),
        "payment_status": s.payment_status.value,
        "items_count": s.items.count(),
        "created_at": s.created_at.isoformat()
    } for s in recent_sales]
    
    # Top products sold by this seller
    from database.models import Product
    products_query = db.query(
        SaleItem.product_id,
        Product.name.label('product_name'),
        func.sum(SaleItem.quantity).label('total_quantity'),
        func.sum(SaleItem.total_price).label('total_revenue'),
        func.count(SaleItem.id).label('times_sold')
    ).join(Sale, Sale.id == SaleItem.sale_id)\
     .join(Product, Product.id == SaleItem.product_id)\
     .filter(
        Sale.seller_id == seller_id,
        Sale.sale_date >= start_date,
        Sale.sale_date <= end_date,
        Sale.is_cancelled == False
    ).group_by(SaleItem.product_id, Product.name)\
     .order_by(func.sum(SaleItem.total_price).desc())\
     .limit(20).all()
    
    top_products = [{
        "product_id": p.product_id,
        "product_name": p.product_name,
        "total_quantity": float(p.total_quantity or 0),
        "total_revenue": float(p.total_revenue or 0),
        "times_sold": p.times_sold
    } for p in products_query]
    
    # Daily breakdown
    daily_query = db.query(
        Sale.sale_date,
        func.count(Sale.id).label('sales_count'),
        func.sum(Sale.total_amount).label('total_amount'),
        func.sum(Sale.paid_amount).label('paid_amount'),
        func.sum(Sale.debt_amount).label('debt_amount')
    ).filter(
        Sale.seller_id == seller_id,
        Sale.sale_date >= start_date,
        Sale.sale_date <= end_date,
        Sale.is_cancelled == False
    ).group_by(Sale.sale_date)\
     .order_by(Sale.sale_date.desc()).all()
    
    daily_breakdown = [{
        "date": d.sale_date.isoformat(),
        "sales_count": d.sales_count,
        "total_amount": float(d.total_amount or 0),
        "paid_amount": float(d.paid_amount or 0),
        "debt_amount": float(d.debt_amount or 0)
    } for d in daily_query]
    
    return {
        "success": True,
        "seller": {
            "id": seller.id,
            "name": f"{seller.first_name} {seller.last_name}",
            "username": seller.username,
            "role": seller.role.display_name if seller.role else "Kassir"
        },
        "summary": {
            "total_sales_count": total_sales_count,
            "total_amount": float(total_amount),
            "total_paid": float(total_paid),
            "total_debt": float(total_debt),
            "total_discount": float(total_discount),
            "average_sale": float(total_amount / total_sales_count) if total_sales_count > 0 else 0,
            "unique_customers": len(customers),
            "anonymous_sales": anonymous_sales,
            "anonymous_amount": float(anonymous_amount)
        },
        "payment_breakdown": payment_breakdown,
        "customers": customers,
        "top_products": top_products,
        "daily_breakdown": daily_breakdown,
        "recent_sales": sales_list,
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat()
        }
    }


@router.get(
    "/sellers-summary",
    summary="Barcha kassirlar xulosasi (JSON)"
)
async def get_sellers_summary(
    start_date: date,
    end_date: date,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get summary statistics for all sellers/cashiers."""
    from sqlalchemy import func
    from database.models import Sale, User as UserModel
    
    # Get all sellers with their stats
    sellers_query = db.query(
        UserModel.id,
        UserModel.first_name,
        UserModel.last_name,
        UserModel.username,
        func.count(Sale.id).label('sales_count'),
        func.sum(Sale.total_amount).label('total_amount'),
        func.sum(Sale.paid_amount).label('total_paid'),
        func.sum(Sale.debt_amount).label('total_debt')
    ).outerjoin(
        Sale, 
        (Sale.seller_id == UserModel.id) & 
        (Sale.sale_date >= start_date) & 
        (Sale.sale_date <= end_date) & 
        (Sale.is_cancelled == False)
    ).filter(
        UserModel.is_active == True
    ).group_by(
        UserModel.id,
        UserModel.first_name,
        UserModel.last_name,
        UserModel.username
    ).order_by(func.sum(Sale.total_amount).desc().nullslast()).all()
    
    sellers = [{
        "id": s.id,
        "name": f"{s.first_name} {s.last_name}",
        "username": s.username,
        "sales_count": s.sales_count or 0,
        "total_amount": float(s.total_amount or 0),
        "total_paid": float(s.total_paid or 0),
        "total_debt": float(s.total_debt or 0)
    } for s in sellers_query]
    
    # Overall totals
    total_sales = sum(s["sales_count"] for s in sellers)
    total_amount = sum(s["total_amount"] for s in sellers)
    total_paid = sum(s["total_paid"] for s in sellers)
    total_debt = sum(s["total_debt"] for s in sellers)
    
    return {
        "success": True,
        "sellers": sellers,
        "totals": {
            "total_sales": total_sales,
            "total_amount": total_amount,
            "total_paid": total_paid,
            "total_debt": total_debt
        },
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat()
        }
    }


@router.get(
    "/excel/seller-stats",
    summary="Kassir statistikasi (Excel)"
)
async def export_seller_stats_excel(
    seller_id: int,
    start_date: date,
    end_date: date,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export seller statistics as Excel file."""
    from sqlalchemy import func
    from decimal import Decimal
    from database.models import Sale, SaleItem, Payment, User as UserModel, Customer, Product
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    import io
    
    # Get seller info
    seller = db.query(UserModel).filter(UserModel.id == seller_id).first()
    if not seller:
        raise HTTPException(status_code=404, detail="Kassir topilmadi")
    
    seller_name = f"{seller.first_name} {seller.last_name}"
    
    # Get all sales by this seller
    sales = db.query(Sale).filter(
        Sale.seller_id == seller_id,
        Sale.sale_date >= start_date,
        Sale.sale_date <= end_date,
        Sale.is_cancelled == False
    ).order_by(Sale.created_at.desc()).all()
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=12)
    money_font = Font(bold=True, color="228B22")
    debt_font = Font(bold=True, color="DC143C")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============ SHEET 1: UMUMIY MA'LUMOT ============
    ws1 = wb.active
    ws1.title = "Umumiy ma'lumot"
    
    # Title
    ws1.merge_cells('A1:D1')
    ws1['A1'] = f"KASSIR HISOBOTI: {seller_name}"
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    ws1['A2'] = f"Davr: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A2:D2')
    
    # Summary stats
    total_amount = sum(s.total_amount for s in sales) if sales else Decimal("0")
    total_paid = sum(s.paid_amount for s in sales) if sales else Decimal("0")
    total_debt = sum(s.debt_amount for s in sales) if sales else Decimal("0")
    total_discount = sum(s.discount_amount for s in sales) if sales else Decimal("0")
    
    row = 4
    stats = [
        ("Jami sotuvlar soni:", len(sales)),
        ("Jami summa:", f"{float(total_amount):,.0f} so'm"),
        ("To'langan summa:", f"{float(total_paid):,.0f} so'm"),
        ("Qarzga sotilgan:", f"{float(total_debt):,.0f} so'm"),
        ("Chegirmalar:", f"{float(total_discount):,.0f} so'm"),
        ("O'rtacha chek:", f"{float(total_amount / len(sales) if sales else 0):,.0f} so'm"),
    ]
    
    for label, value in stats:
        ws1[f'A{row}'] = label
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = value
        if "Qarz" in label:
            ws1[f'B{row}'].font = debt_font
        elif "summa" in label.lower() or "chek" in label.lower():
            ws1[f'B{row}'].font = money_font
        row += 1
    
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20
    
    # ============ SHEET 2: SOTUVLAR RO'YXATI ============
    ws2 = wb.create_sheet("Sotuvlar")
    
    headers = ["#", "Sana", "Chek №", "Mijoz", "Summa", "To'langan", "Qarz", "To'lov turi", "Mahsulotlar"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    for idx, sale in enumerate(sales, 1):
        payment_type_str = ""
        if sale.payment_type:
            pt_map = {"CASH": "Naqd", "CARD": "Karta", "TRANSFER": "O'tkazma", "DEBT": "Qarz"}
            payment_type_str = pt_map.get(sale.payment_type.value if hasattr(sale.payment_type, 'value') else str(sale.payment_type), "")
        
        row_data = [
            idx,
            sale.created_at.strftime('%d.%m.%Y %H:%M'),
            sale.sale_number,
            sale.customer.name if sale.customer else "Oddiy xaridor",
            float(sale.total_amount),
            float(sale.paid_amount),
            float(sale.debt_amount),
            payment_type_str,
            sale.items.count()
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=idx + 1, column=col, value=value)
            cell.border = thin_border
            if col in [5, 6, 7]:  # Money columns
                cell.number_format = '#,##0'
                if col == 7 and value > 0:  # Debt
                    cell.font = debt_font
    
    # Set column widths
    col_widths = [5, 18, 12, 25, 15, 15, 15, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width
    
    # ============ SHEET 3: MIJOZLAR ============
    ws3 = wb.create_sheet("Mijozlar")
    
    # Get customers
    customers_data = db.query(
        Customer.id,
        Customer.name,
        Customer.phone,
        Customer.company_name,
        func.count(Sale.id).label('sales_count'),
        func.sum(Sale.total_amount).label('total_amount'),
        func.sum(Sale.debt_amount).label('total_debt')
    ).join(Sale, Sale.customer_id == Customer.id)\
     .filter(
        Sale.seller_id == seller_id,
        Sale.sale_date >= start_date,
        Sale.sale_date <= end_date,
        Sale.is_cancelled == False
    ).group_by(Customer.id, Customer.name, Customer.phone, Customer.company_name)\
     .order_by(func.sum(Sale.total_amount).desc()).all()
    
    headers = ["#", "Mijoz", "Telefon", "Kompaniya", "Xaridlar", "Jami summa", "Qarz"]
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for idx, cust in enumerate(customers_data, 1):
        row_data = [
            idx,
            cust.name,
            cust.phone,
            cust.company_name or "",
            cust.sales_count,
            float(cust.total_amount or 0),
            float(cust.total_debt or 0)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws3.cell(row=idx + 1, column=col, value=value)
            cell.border = thin_border
            if col in [6, 7]:
                cell.number_format = '#,##0'
                if col == 7 and value > 0:
                    cell.font = debt_font
    
    col_widths = [5, 25, 15, 20, 10, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = width
    
    # ============ SHEET 4: QARZDORLAR ============
    ws4 = wb.create_sheet("Qarzdorlar")
    
    # Get debt sales
    debt_sales = [s for s in sales if s.debt_amount > 0]
    
    headers = ["#", "Sana", "Chek №", "Mijoz", "Telefon", "Jami summa", "Qarz summa"]
    for col, header in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.border = thin_border
    
    for idx, sale in enumerate(debt_sales, 1):
        row_data = [
            idx,
            sale.created_at.strftime('%d.%m.%Y %H:%M'),
            sale.sale_number,
            sale.customer.name if sale.customer else "Oddiy xaridor",
            sale.customer.phone if sale.customer else "",
            float(sale.total_amount),
            float(sale.debt_amount)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws4.cell(row=idx + 1, column=col, value=value)
            cell.border = thin_border
            if col in [6, 7]:
                cell.number_format = '#,##0'
                if col == 7:
                    cell.font = debt_font
    
    col_widths = [5, 18, 12, 25, 15, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws4.column_dimensions[get_column_letter(i)].width = width
    
    # ============ SHEET 5: TOP MAHSULOTLAR ============
    ws5 = wb.create_sheet("Top mahsulotlar")
    
    products_data = db.query(
        SaleItem.product_id,
        Product.name.label('product_name'),
        func.sum(SaleItem.quantity).label('total_quantity'),
        func.sum(SaleItem.total_price).label('total_revenue'),
        func.count(SaleItem.id).label('times_sold')
    ).join(Sale, Sale.id == SaleItem.sale_id)\
     .join(Product, Product.id == SaleItem.product_id)\
     .filter(
        Sale.seller_id == seller_id,
        Sale.sale_date >= start_date,
        Sale.sale_date <= end_date,
        Sale.is_cancelled == False
    ).group_by(SaleItem.product_id, Product.name)\
     .order_by(func.sum(SaleItem.total_price).desc())\
     .limit(50).all()
    
    headers = ["#", "Mahsulot", "Sotilgan soni", "Necha marta", "Jami summa"]
    for col, header in enumerate(headers, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.border = thin_border
    
    for idx, prod in enumerate(products_data, 1):
        row_data = [
            idx,
            prod.product_name,
            float(prod.total_quantity or 0),
            prod.times_sold,
            float(prod.total_revenue or 0)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws5.cell(row=idx + 1, column=col, value=value)
            cell.border = thin_border
            if col == 5:
                cell.number_format = '#,##0'
                cell.font = money_font
    
    col_widths = [5, 40, 15, 12, 18]
    for i, width in enumerate(col_widths, 1):
        ws5.column_dimensions[get_column_letter(i)].width = width
    
    # ============ SHEET 6: KUNLIK STATISTIKA ============
    ws6 = wb.create_sheet("Kunlik statistika")
    
    daily_data = db.query(
        Sale.sale_date,
        func.count(Sale.id).label('sales_count'),
        func.sum(Sale.total_amount).label('total_amount'),
        func.sum(Sale.paid_amount).label('paid_amount'),
        func.sum(Sale.debt_amount).label('debt_amount')
    ).filter(
        Sale.seller_id == seller_id,
        Sale.sale_date >= start_date,
        Sale.sale_date <= end_date,
        Sale.is_cancelled == False
    ).group_by(Sale.sale_date).order_by(Sale.sale_date.desc()).all()
    
    headers = ["Sana", "Sotuvlar", "Jami summa", "To'langan", "Qarz"]
    for col, header in enumerate(headers, 1):
        cell = ws6.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        cell.border = thin_border
    
    for idx, day in enumerate(daily_data, 1):
        row_data = [
            day.sale_date.strftime('%d.%m.%Y'),
            day.sales_count,
            float(day.total_amount or 0),
            float(day.paid_amount or 0),
            float(day.debt_amount or 0)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws6.cell(row=idx + 1, column=col, value=value)
            cell.border = thin_border
            if col in [3, 4, 5]:
                cell.number_format = '#,##0'
                if col == 5 and value > 0:
                    cell.font = debt_font
    
    col_widths = [15, 12, 18, 18, 18]
    for i, width in enumerate(col_widths, 1):
        ws6.column_dimensions[get_column_letter(i)].width = width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"kassir_{seller.username}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )