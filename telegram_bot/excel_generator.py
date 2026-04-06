"""
Excel Generator for Telegram Bot Notifications
Creates professional Excel files for purchase and payment notifications.
"""
import os
import io
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class NotificationExcelGenerator:
    """Generates Excel files for customer notifications."""
    
    def __init__(self, company_name: str = "Vegas"):
        self.company_name = company_name
    
    def _format_money(self, amount: Any) -> str:
        """Format money amount."""
        if amount is None:
            return "0"
        try:
            num = float(amount)
            if num >= 1_000_000:
                return f"{num:,.0f}".replace(",", " ")
            return f"{num:,.2f}".replace(",", " ")
        except (ValueError, TypeError):
            return str(amount)
    
    def _setup_styles(self):
        """Setup common styles."""
        self.header_font = Font(bold=True, size=14)
        self.title_font = Font(bold=True, size=12)
        self.normal_font = Font(size=11)
        self.money_font = Font(size=11, bold=True)
        
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font_white = Font(bold=True, size=11, color="FFFFFF")
        
        self.success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    def generate_purchase_notification(
        self,
        customer_name: str,
        customer_phone: str,
        sale_number: str,
        sale_date: datetime,
        items: List[Dict[str, Any]],
        total_amount: float,
        paid_amount: float,
        debt_amount: float,
        operator_name: str = "Kassir"
    ) -> io.BytesIO:
        """
        Generate Excel file for purchase notification.
        
        Args:
            customer_name: Customer's name
            customer_phone: Customer's phone
            sale_number: Sale number/ID
            sale_date: Date of sale
            items: List of purchased items
            total_amount: Total sale amount
            paid_amount: Amount paid
            debt_amount: Remaining debt
            operator_name: Name of the operator/cashier
        
        Returns:
            BytesIO object containing the Excel file
        """
        self._setup_styles()
        wb = Workbook()
        ws = wb.active
        ws.title = "Harid ma'lumotlari"
        
        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = f"📦 HARID CHEKI - {self.company_name}"
        ws['A1'].font = self.header_font
        ws['A1'].alignment = self.center_align
        
        # Customer info
        info_data = [
            ('Mijoz:', customer_name),
            ('Telefon:', customer_phone),
            ('Chek raqami:', sale_number),
            ('Sana:', sale_date.strftime('%d.%m.%Y %H:%M') if isinstance(sale_date, datetime) else str(sale_date)),
            ('Kassir:', operator_name),
        ]
        
        row = 3
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.title_font
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.normal_font
            row += 1
        
        # Items table header
        row += 1
        headers = ['№', 'Tovar nomi', 'Miqdor', 'Birlik narx', 'Chegirma', 'Jami']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        
        # Items data
        row += 1
        for idx, item in enumerate(items, 1):
            ws.cell(row=row, column=1, value=idx).border = self.thin_border
            ws.cell(row=row, column=2, value=item.get('product_name', '')).border = self.thin_border
            
            qty_cell = ws.cell(row=row, column=3, value=f"{item.get('quantity', 0)} {item.get('uom_symbol', '')}")
            qty_cell.border = self.thin_border
            qty_cell.alignment = self.center_align
            
            price_cell = ws.cell(row=row, column=4, value=self._format_money(item.get('unit_price', 0)))
            price_cell.border = self.thin_border
            price_cell.alignment = self.right_align
            
            discount_cell = ws.cell(row=row, column=5, value=self._format_money(item.get('discount_amount', 0)))
            discount_cell.border = self.thin_border
            discount_cell.alignment = self.right_align
            
            total_cell = ws.cell(row=row, column=6, value=self._format_money(item.get('total_price', 0)))
            total_cell.border = self.thin_border
            total_cell.alignment = self.right_align
            total_cell.font = self.money_font
            
            row += 1
        
        # Summary
        row += 1
        summary_data = [
            ('JAMI SUMMA:', total_amount, None),
            ("TO'LANGAN:", paid_amount, self.success_fill if paid_amount > 0 else None),
            ('QARZ QOLDI:', debt_amount, self.warning_fill if debt_amount > 0 else self.success_fill),
        ]
        
        for label, amount, fill in summary_data:
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.title_font
            ws[f'A{row}'].alignment = self.right_align
            
            ws[f'F{row}'] = self._format_money(amount)
            ws[f'F{row}'].font = self.money_font
            ws[f'F{row}'].alignment = self.right_align
            ws[f'F{row}'].border = self.thin_border
            if fill:
                ws[f'F{row}'].fill = fill
            row += 1
        
        # Footer
        row += 2
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"✅ Xaridingiz uchun rahmat! - {self.company_name}"
        ws[f'A{row}'].font = Font(italic=True, size=11)
        ws[f'A{row}'].alignment = self.center_align
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def generate_payment_notification(
        self,
        customer_name: str,
        customer_phone: str,
        payment_date: datetime,
        payment_amount: float,
        payment_type: str,
        previous_debt: float,
        current_debt: float,
        operator_name: str = "Kassir"
    ) -> io.BytesIO:
        """
        Generate Excel file for payment notification.
        
        Args:
            customer_name: Customer's name
            customer_phone: Customer's phone
            payment_date: Date of payment
            payment_amount: Amount paid
            payment_type: Payment type (CASH, CARD, etc.)
            previous_debt: Debt before payment
            current_debt: Debt after payment
            operator_name: Name of the operator
        
        Returns:
            BytesIO object containing the Excel file
        """
        self._setup_styles()
        wb = Workbook()
        ws = wb.active
        ws.title = "To'lov ma'lumotlari"
        
        # Header
        ws.merge_cells('A1:D1')
        ws['A1'] = f"💰 TO'LOV KVITANSIYASI - {self.company_name}"
        ws['A1'].font = self.header_font
        ws['A1'].alignment = self.center_align
        
        # Payment type translation
        payment_type_labels = {
            'CASH': 'Naqd pul',
            'CARD': 'Plastik karta',
            'TRANSFER': "Bank o'tkazmasi",
            'MIXED': 'Aralash'
        }
        
        # Payment info
        info_data = [
            ('Mijoz:', customer_name),
            ('Telefon:', customer_phone),
            ('Sana:', payment_date.strftime('%d.%m.%Y %H:%M') if isinstance(payment_date, datetime) else str(payment_date)),
            ("To'lov turi:", payment_type_labels.get(payment_type, payment_type)),
            ('Qabul qildi:', operator_name),
        ]
        
        row = 3
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.title_font
            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.normal_font
            row += 1
        
        # Payment summary table
        row += 1
        headers = ['Tavsif', 'Summa']
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = headers[0]
        ws[f'A{row}'].font = self.header_font_white
        ws[f'A{row}'].fill = self.header_fill
        ws[f'A{row}'].border = self.thin_border
        
        ws[f'D{row}'] = headers[1]
        ws[f'D{row}'].font = self.header_font_white
        ws[f'D{row}'].fill = self.header_fill
        ws[f'D{row}'].border = self.thin_border
        ws[f'D{row}'].alignment = self.right_align
        
        row += 1
        payment_data = [
            ("Oldingi qarz:", previous_debt, None),
            ("To'langan summa:", payment_amount, self.success_fill),
            ("Joriy qarz:", current_debt, self.warning_fill if current_debt > 0 else self.success_fill),
        ]
        
        for label, amount, fill in payment_data:
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.normal_font
            ws[f'A{row}'].border = self.thin_border
            
            ws[f'D{row}'] = self._format_money(amount)
            ws[f'D{row}'].font = self.money_font
            ws[f'D{row}'].alignment = self.right_align
            ws[f'D{row}'].border = self.thin_border
            if fill:
                ws[f'D{row}'].fill = fill
            row += 1
        
        # Status
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        if current_debt <= 0:
            ws[f'A{row}'] = "✅ Qarz to'liq to'landi!"
            ws[f'A{row}'].fill = self.success_fill
        else:
            ws[f'A{row}'] = f"⚠️ Qolgan qarz: {self._format_money(current_debt)} so'm"
            ws[f'A{row}'].fill = self.warning_fill
        ws[f'A{row}'].font = self.title_font
        ws[f'A{row}'].alignment = self.center_align
        
        # Footer
        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = f"Rahmat! - {self.company_name}"
        ws[f'A{row}'].font = Font(italic=True, size=11)
        ws[f'A{row}'].alignment = self.center_align
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


# Create global instance
excel_generator = NotificationExcelGenerator()
