"""
Excel Report Generator Service.
Generates Excel reports for sales, stock, customers, etc.
"""

import io
from datetime import datetime, date
from decimal import Decimal
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import func, case as sa_case

from database.models import (
    Sale, SaleItem, Product, Customer, Stock, StockMovement,
    Payment, Warehouse, Category, Expense, ExpenseCategory
)
from utils.helpers import get_tashkent_datetime_str, get_tashkent_date_str


class ExcelReportGenerator:
    """Excel report generator with professional styling."""
    
    # Styles
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    TITLE_FONT = Font(bold=True, size=14)
    SUBTITLE_FONT = Font(bold=True, size=11)
    
    CURRENCY_FORMAT = '#,##0'
    DATE_FORMAT = 'DD.MM.YYYY'
    DATETIME_FORMAT = 'DD.MM.YYYY HH:MM'
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, db: Session):
        self.db = db
    
    def _create_workbook(self) -> Workbook:
        """Create new workbook with default settings."""
        wb = Workbook()
        return wb
    
    def _set_column_widths(self, ws, widths: List[int]):
        """Set column widths."""
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def _add_header_row(self, ws, row: int, headers: List[str]):
        """Add styled header row."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
    
    def _add_title(self, ws, title: str, subtitle: str = None):
        """Add report title."""
        ws.cell(row=1, column=1, value=title).font = self.TITLE_FONT
        if subtitle:
            ws.cell(row=2, column=1, value=subtitle).font = self.SUBTITLE_FONT
    
    def generate_sales_report(
        self,
        start_date: date,
        end_date: date,
        warehouse_id: int = None,
        seller_id: int = None
    ) -> bytes:
        """
        Generate sales report Excel file.
        
        Returns: Excel file as bytes
        """
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "Sotuvlar hisoboti"
        
        # Title
        self._add_title(
            ws,
            "SOTUVLAR HISOBOTI",
            f"Davr: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
        )
        
        # Query sales
        query = self.db.query(Sale).filter(
            Sale.sale_date >= start_date,
            Sale.sale_date <= end_date,
            Sale.is_cancelled == False
        )
        
        if warehouse_id:
            query = query.filter(Sale.warehouse_id == warehouse_id)
        if seller_id:
            query = query.filter(Sale.seller_id == seller_id)
        
        sales = query.order_by(Sale.sale_date, Sale.id).all()
        
        # Headers
        headers = [
            "№", "Sana", "Sotuv №", "Mijoz", "Sotuvchi",
            "Jami summa", "Chegirma", "Yakuniy", "To'langan", "Qarz", "Status"
        ]
        self._add_header_row(ws, 4, headers)
        self._set_column_widths(ws, [5, 12, 15, 25, 20, 15, 12, 15, 15, 15, 12])
        
        # Data rows
        row = 5
        totals = {
            'subtotal': Decimal('0'),
            'discount': Decimal('0'),
            'total': Decimal('0'),
            'paid': Decimal('0'),
            'debt': Decimal('0')
        }
        
        for i, sale in enumerate(sales, 1):
            ws.cell(row=row, column=1, value=i).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=sale.sale_date).border = self.THIN_BORDER
            ws.cell(row=row, column=2).number_format = self.DATE_FORMAT
            ws.cell(row=row, column=3, value=sale.sale_number).border = self.THIN_BORDER
            ws.cell(row=row, column=4, value=sale.customer.name if sale.customer else "Noma'lum").border = self.THIN_BORDER
            ws.cell(row=row, column=5, value=f"{sale.seller.first_name} {sale.seller.last_name}").border = self.THIN_BORDER
            ws.cell(row=row, column=6, value=float(sale.subtotal)).border = self.THIN_BORDER
            ws.cell(row=row, column=6).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=7, value=float(sale.discount_amount)).border = self.THIN_BORDER
            ws.cell(row=row, column=7).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=8, value=float(sale.total_amount)).border = self.THIN_BORDER
            ws.cell(row=row, column=8).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=9, value=float(sale.paid_amount)).border = self.THIN_BORDER
            ws.cell(row=row, column=9).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=10, value=float(sale.debt_amount)).border = self.THIN_BORDER
            ws.cell(row=row, column=10).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=11, value=sale.payment_status.value).border = self.THIN_BORDER
            
            totals['subtotal'] += sale.subtotal
            totals['discount'] += sale.discount_amount
            totals['total'] += sale.total_amount
            totals['paid'] += sale.paid_amount
            totals['debt'] += sale.debt_amount
            
            row += 1
        
        # Totals row
        row += 1
        ws.cell(row=row, column=5, value="JAMI:").font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(totals['subtotal'])).font = Font(bold=True)
        ws.cell(row=row, column=6).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=7, value=float(totals['discount'])).font = Font(bold=True)
        ws.cell(row=row, column=7).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=8, value=float(totals['total'])).font = Font(bold=True)
        ws.cell(row=row, column=8).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=9, value=float(totals['paid'])).font = Font(bold=True)
        ws.cell(row=row, column=9).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=10, value=float(totals['debt'])).font = Font(bold=True)
        ws.cell(row=row, column=10).number_format = self.CURRENCY_FORMAT
        
        # Summary
        row += 2
        ws.cell(row=row, column=1, value=f"Jami sotuvlar soni: {len(sales)}")

        # ===== FOYDA VA CHIQIMLAR =====
        row += 2
        ws.cell(row=row, column=1, value="FOYDA HISOBOTI").font = Font(bold=True, size=12, color="228B22")
        row += 1

        # Gross profit from SaleItems
        from sqlalchemy import case as sa_case
        _eff_cost = sa_case((SaleItem.unit_cost > 0, SaleItem.unit_cost), else_=Product.cost_price)
        profit_data = self.db.query(
            func.coalesce(func.sum(SaleItem.total_price), 0).label('revenue'),
            func.coalesce(func.sum(_eff_cost * SaleItem.base_quantity), 0).label('cost')
        ).join(Sale, Sale.id == SaleItem.sale_id).join(Product, Product.id == SaleItem.product_id).filter(
            Sale.sale_date >= start_date,
            Sale.sale_date <= end_date,
            Sale.is_cancelled == False
        )
        if warehouse_id:
            profit_data = profit_data.filter(Sale.warehouse_id == warehouse_id)
        profit_result = profit_data.first()

        gross_revenue = float(profit_result.revenue or 0) if profit_result else 0
        gross_cost = float(profit_result.cost or 0) if profit_result else 0
        gross_profit = gross_revenue - gross_cost

        ws.cell(row=row, column=1, value="Sotuv summasi:")
        ws.cell(row=row, column=2, value=gross_revenue).number_format = self.CURRENCY_FORMAT
        row += 1
        ws.cell(row=row, column=1, value="Kelish narxi (tan narxi):")
        ws.cell(row=row, column=2, value=gross_cost).number_format = self.CURRENCY_FORMAT
        row += 1
        ws.cell(row=row, column=1, value="Yalpi foyda:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=gross_profit).font = Font(bold=True, color="228B22")
        ws.cell(row=row, column=2).number_format = self.CURRENCY_FORMAT
        row += 2

        # Expenses
        ws.cell(row=row, column=1, value="CHIQIMLAR").font = Font(bold=True, size=12, color="DC143C")
        row += 1

        expenses = self.db.query(Expense).filter(
            Expense.expense_date >= start_date,
            Expense.expense_date <= end_date
        ).all()
        total_expenses = sum(float(e.amount or 0) for e in expenses)

        # By category
        exp_by_cat = self.db.query(
            ExpenseCategory.name,
            func.coalesce(func.sum(Expense.amount), 0).label('total')
        ).outerjoin(Expense, (Expense.category_id == ExpenseCategory.id) &
                    (Expense.expense_date >= start_date) &
                    (Expense.expense_date <= end_date))\
         .filter(ExpenseCategory.is_active == True)\
         .group_by(ExpenseCategory.id, ExpenseCategory.name)\
         .having(func.sum(Expense.amount) > 0).all()

        for cat_name, cat_total in exp_by_cat:
            ws.cell(row=row, column=1, value=f"  {cat_name}:")
            ws.cell(row=row, column=2, value=float(cat_total or 0)).number_format = self.CURRENCY_FORMAT
            row += 1

        ws.cell(row=row, column=1, value="Jami chiqimlar:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_expenses).font = Font(bold=True, color="DC143C")
        ws.cell(row=row, column=2).number_format = self.CURRENCY_FORMAT
        row += 2

        # Net profit
        net_profit = gross_profit - total_expenses
        ws.cell(row=row, column=1, value="SOF FOYDA:").font = Font(bold=True, size=14)
        profit_cell = ws.cell(row=row, column=2, value=net_profit)
        profit_cell.font = Font(bold=True, size=14, color="228B22" if net_profit >= 0 else "DC143C")
        profit_cell.number_format = self.CURRENCY_FORMAT
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def generate_stock_report(self, warehouse_id: int = None) -> bytes:
        """Generate stock/inventory report."""
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "Qoldiqlar hisoboti"
        
        # Title
        self._add_title(
            ws,
            "OMBOR QOLDIQLARI HISOBOTI",
            f"Sana: {get_tashkent_datetime_str()}"
        )
        
        # Query stock
        query = self.db.query(Stock).join(Product).filter(Product.is_deleted == False)
        
        if warehouse_id:
            query = query.filter(Stock.warehouse_id == warehouse_id)
        
        stocks = query.order_by(Product.name).all()
        
        # Headers
        headers = [
            "№", "Tovar nomi", "Artikul", "Kategoriya", "Ombor",
            "Miqdor", "O'lchov", "O'rtacha narx", "Jami qiymat", "Min. qoldiq", "Holat"
        ]
        self._add_header_row(ws, 4, headers)
        self._set_column_widths(ws, [5, 30, 15, 20, 15, 12, 8, 15, 18, 12, 12])
        
        # Data rows
        row = 5
        total_value = Decimal('0')
        below_min_count = 0
        
        for i, stock in enumerate(stocks, 1):
            product = stock.product
            value = stock.quantity * stock.average_cost
            is_below_min = stock.quantity < product.min_stock_level
            
            if is_below_min:
                below_min_count += 1
            
            ws.cell(row=row, column=1, value=i).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=product.name).border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=product.article or "-").border = self.THIN_BORDER
            ws.cell(row=row, column=4, value=product.category.name if product.category else "-").border = self.THIN_BORDER
            ws.cell(row=row, column=5, value=stock.warehouse.name).border = self.THIN_BORDER
            ws.cell(row=row, column=6, value=float(stock.quantity)).border = self.THIN_BORDER
            ws.cell(row=row, column=7, value=product.base_uom.symbol).border = self.THIN_BORDER
            ws.cell(row=row, column=8, value=float(stock.average_cost)).border = self.THIN_BORDER
            ws.cell(row=row, column=8).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=9, value=float(value)).border = self.THIN_BORDER
            ws.cell(row=row, column=9).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=10, value=float(product.min_stock_level)).border = self.THIN_BORDER
            
            status_cell = ws.cell(row=row, column=11, value="Kam!" if is_below_min else "OK")
            status_cell.border = self.THIN_BORDER
            if is_below_min:
                status_cell.font = Font(color="FF0000", bold=True)
            
            total_value += value
            row += 1
        
        # Summary
        row += 1
        ws.cell(row=row, column=8, value="JAMI QIYMAT:").font = Font(bold=True)
        ws.cell(row=row, column=9, value=float(total_value)).font = Font(bold=True)
        ws.cell(row=row, column=9).number_format = self.CURRENCY_FORMAT
        
        row += 2
        ws.cell(row=row, column=1, value=f"Jami tovarlar: {len(stocks)}")
        ws.cell(row=row+1, column=1, value=f"Kam qoldiqli: {below_min_count}")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def generate_debtors_report(self) -> bytes:
        """Generate customer debtors report."""
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "Qarzdorlar"
        
        self._add_title(
            ws,
            "QARZDORLAR HISOBOTI",
            f"Sana: {get_tashkent_datetime_str()}"
        )
        
        # Query debtors
        debtors = self.db.query(Customer).filter(
            Customer.is_deleted == False,
            Customer.current_debt > 0
        ).order_by(Customer.current_debt.desc()).all()
        
        headers = [
            "№", "Mijoz", "Telefon", "Kompaniya", "Qarz summasi",
            "Kredit limit", "Oxirgi xarid", "Manager"
        ]
        self._add_header_row(ws, 4, headers)
        self._set_column_widths(ws, [5, 25, 15, 20, 18, 15, 12, 20])
        
        row = 5
        total_debt = Decimal('0')
        
        for i, customer in enumerate(debtors, 1):
            ws.cell(row=row, column=1, value=i).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=customer.name).border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=customer.phone).border = self.THIN_BORDER
            ws.cell(row=row, column=4, value=customer.company_name or "-").border = self.THIN_BORDER
            ws.cell(row=row, column=5, value=float(customer.current_debt)).border = self.THIN_BORDER
            ws.cell(row=row, column=5).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=6, value=float(customer.credit_limit)).border = self.THIN_BORDER
            ws.cell(row=row, column=6).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=7, value=customer.last_purchase_date).border = self.THIN_BORDER
            if customer.last_purchase_date:
                ws.cell(row=row, column=7).number_format = self.DATE_FORMAT
            ws.cell(row=row, column=8, value=f"{customer.manager.first_name} {customer.manager.last_name}" if customer.manager else "-").border = self.THIN_BORDER
            
            total_debt += customer.current_debt
            row += 1
        
        row += 1
        ws.cell(row=row, column=4, value="JAMI QARZ:").font = Font(bold=True)
        ws.cell(row=row, column=5, value=float(total_debt)).font = Font(bold=True)
        ws.cell(row=row, column=5).number_format = self.CURRENCY_FORMAT
        
        row += 2
        ws.cell(row=row, column=1, value=f"Jami qarzdorlar: {len(debtors)}")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def generate_daily_report(self, report_date: date, warehouse_id: int = None) -> bytes:
        """Generate daily summary report."""
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "Kunlik hisobot"
        
        self._add_title(
            ws,
            "KUNLIK HISOBOT",
            f"Sana: {report_date.strftime('%d.%m.%Y')}"
        )
        
        # Query data
        sales_query = self.db.query(Sale).filter(
            Sale.sale_date == report_date,
            Sale.is_cancelled == False
        )
        if warehouse_id:
            sales_query = sales_query.filter(Sale.warehouse_id == warehouse_id)
        
        sales = sales_query.all()
        
        # Sales summary
        row = 4
        ws.cell(row=row, column=1, value="SOTUVLAR").font = Font(bold=True, size=12)
        row += 1
        
        total_sales = len(sales)
        total_amount = sum(s.total_amount for s in sales)
        total_paid = sum(s.paid_amount for s in sales)
        total_debt = sum(s.debt_amount for s in sales)
        total_discount = sum(s.discount_amount for s in sales)
        
        summary_data = [
            ("Sotuvlar soni:", total_sales),
            ("Jami summa:", f"{total_amount:,.0f} so'm"),
            ("Chegirmalar:", f"{total_discount:,.0f} so'm"),
            ("Naqd to'lovlar:", f"{total_paid:,.0f} so'm"),
            ("Qarzga:", f"{total_debt:,.0f} so'm"),
        ]
        
        for label, value in summary_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Payment breakdown
        row += 1
        ws.cell(row=row, column=1, value="TO'LOV TURLARI").font = Font(bold=True, size=12)
        row += 1
        
        payments = self.db.query(
            Payment.payment_type,
            func.sum(Payment.amount)
        ).filter(
            Payment.payment_date == report_date,
            Payment.is_cancelled == False
        ).group_by(Payment.payment_type).all()
        
        for payment_type, amount in payments:
            ws.cell(row=row, column=1, value=payment_type.value)
            ws.cell(row=row, column=2, value=f"{amount:,.0f} so'm")
            row += 1
        
        # Top products
        row += 1
        ws.cell(row=row, column=1, value="ENG KO'P SOTILGAN TOVARLAR").font = Font(bold=True, size=12)
        row += 1
        
        top_products = self.db.query(
            Product.name,
            func.sum(SaleItem.base_quantity).label('total_qty'),
            func.sum(SaleItem.total_price).label('total_amount')
        ).join(SaleItem).join(Sale).filter(
            Sale.sale_date == report_date,
            Sale.is_cancelled == False
        ).group_by(Product.id, Product.name).order_by(
            func.sum(SaleItem.total_price).desc()
        ).limit(10).all()
        
        headers = ["Tovar", "Miqdor", "Summa"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)
        row += 1
        
        for name, qty, amount in top_products:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=float(qty))
            ws.cell(row=row, column=3, value=f"{amount:,.0f} so'm")
            row += 1

        # ===== FOYDA VA CHIQIMLAR =====
        row += 1
        ws.cell(row=row, column=1, value="FOYDA VA CHIQIMLAR").font = Font(bold=True, size=12)
        row += 1

        # Gross profit
        from sqlalchemy import case as sa_case
        _eff_cost2 = sa_case((SaleItem.unit_cost > 0, SaleItem.unit_cost), else_=Product.cost_price)
        profit_data = self.db.query(
            func.coalesce(func.sum(SaleItem.total_price), 0).label('revenue'),
            func.coalesce(func.sum(_eff_cost2 * SaleItem.base_quantity), 0).label('cost')
        ).join(Sale, Sale.id == SaleItem.sale_id).join(Product, Product.id == SaleItem.product_id).filter(
            Sale.sale_date == report_date,
            Sale.is_cancelled == False
        ).first()

        gross_revenue = float(profit_data.revenue or 0) if profit_data else 0
        gross_cost = float(profit_data.cost or 0) if profit_data else 0
        gross_profit = gross_revenue - gross_cost

        daily_summary = [
            ("Sotuv summasi:", f"{gross_revenue:,.0f} so'm"),
            ("Kelish narxi:", f"{gross_cost:,.0f} so'm"),
            ("Yalpi foyda:", f"{gross_profit:,.0f} so'm"),
        ]
        for label, value in daily_summary:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Expenses
        row += 1
        ws.cell(row=row, column=1, value="Chiqimlar:").font = Font(bold=True)
        row += 1

        day_expenses = self.db.query(Expense).filter(
            Expense.expense_date == report_date
        ).all()
        total_day_expenses = sum(float(e.amount or 0) for e in day_expenses)

        for exp in day_expenses:
            cat_name = exp.category.name if exp.category else "Boshqa"
            ws.cell(row=row, column=1, value=f"  {cat_name}: {exp.description}")
            ws.cell(row=row, column=2, value=f"{float(exp.amount or 0):,.0f} so'm")
            row += 1

        ws.cell(row=row, column=1, value="Jami chiqim:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"{total_day_expenses:,.0f} so'm").font = Font(bold=True, color="DC143C")
        row += 1

        # Net profit
        net_profit = gross_profit - total_day_expenses
        row += 1
        ws.cell(row=row, column=1, value="SOF FOYDA:").font = Font(bold=True, size=14)
        ws.cell(row=row, column=2, value=f"{net_profit:,.0f} so'm").font = Font(
            bold=True, size=14, color="228B22" if net_profit >= 0 else "DC143C"
        )
        
        self._set_column_widths(ws, [30, 15, 20])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def generate_products_price_list(self, category_id: int = None) -> bytes:
        """Generate product price list."""
        wb = self._create_workbook()
        ws = wb.active
        ws.title = "Narxlar"
        
        self._add_title(ws, "TOVARLAR NARX RO'YXATI", f"Sana: {get_tashkent_date_str()}")
        
        query = self.db.query(Product).filter(
            Product.is_deleted == False,
            Product.is_active == True
        )
        
        if category_id:
            query = query.filter(Product.category_id == category_id)
        
        products = query.order_by(Product.category_id, Product.name).all()
        
        headers = ["№", "Artikul", "Tovar nomi", "Kategoriya", "O'lchov", "Narx", "VIP narx"]
        self._add_header_row(ws, 4, headers)
        self._set_column_widths(ws, [5, 15, 35, 20, 10, 15, 15])
        
        row = 5
        for i, product in enumerate(products, 1):
            ws.cell(row=row, column=1, value=i).border = self.THIN_BORDER
            ws.cell(row=row, column=2, value=product.article or "-").border = self.THIN_BORDER
            ws.cell(row=row, column=3, value=product.name).border = self.THIN_BORDER
            ws.cell(row=row, column=4, value=product.category.name if product.category else "-").border = self.THIN_BORDER
            ws.cell(row=row, column=5, value=product.base_uom.symbol).border = self.THIN_BORDER
            ws.cell(row=row, column=6, value=float(product.sale_price)).border = self.THIN_BORDER
            ws.cell(row=row, column=6).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=7, value=float(product.vip_price) if product.vip_price else "-").border = self.THIN_BORDER
            if product.vip_price:
                ws.cell(row=row, column=7).number_format = self.CURRENCY_FORMAT
            row += 1
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
